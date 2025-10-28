import tkinter as tk
from tkinter import ttk, filedialog, simpledialog, messagebox
import openpyxl
import xlrd
from openpyxl.utils.exceptions import InvalidFileException
import pandas as pd

class ExcelMergerApp(tk.Tk):
    def __init__(self):
        super().__init__()
        self.title("Excel Sheet Merger")
        self.geometry("800x600")

        self.file_passwords = {} # To store passwords for encrypted files
        self.selected_sheets = {} # To store selected sheets for each file: {file_path: [sheet_name1, sheet_name2]}

        # 메인 프레임
        main_frame = ttk.Frame(self, padding="10")
        main_frame.grid(row=0, column=0, sticky=(tk.W, tk.E, tk.N, tk.S))
        self.columnconfigure(0, weight=1)
        self.rowconfigure(0, weight=1)

        # 파일 프레임
        file_frame = ttk.LabelFrame(main_frame, text="파일 목록")
        file_frame.grid(row=0, column=0, columnspan=3, padx=5, pady=5, sticky=(tk.W, tk.E, tk.N, tk.S))
        main_frame.columnconfigure(0, weight=1)
        main_frame.rowconfigure(0, weight=1)

        # 버튼 프레임
        button_frame = ttk.Frame(file_frame)
        button_frame.grid(row=0, column=1, padx=5, pady=5, sticky=tk.N)

        # 파일 선택 버튼
        self.select_files_button = ttk.Button(file_frame, text="파일 추가", command=self.select_files)
        self.select_files_button.grid(row=0, column=0, padx=5, pady=5, sticky=(tk.W, tk.N))

        # 파일 제외 버튼
        self.exclude_files_button = ttk.Button(button_frame, text="파일 제외", command=self.exclude_files)
        self.exclude_files_button.pack(fill=tk.X, padx=5, pady=5)

        # 위로 이동 버튼
        self.up_button = ttk.Button(button_frame, text="▲", command=self.move_up)
        self.up_button.pack(fill=tk.X, padx=5, pady=5)

        # 아래로 이동 버튼
        self.down_button = ttk.Button(button_frame, text="▼", command=self.move_down)
        self.down_button.pack(fill=tk.X, padx=5, pady=5)

        # 파일 목록 리스트박스
        self.file_listbox = tk.Listbox(file_frame, selectmode=tk.EXTENDED)
        self.file_listbox.grid(row=1, column=0, padx=5, pady=5, sticky=(tk.W, tk.E, tk.N, tk.S))
        file_frame.columnconfigure(0, weight=1)
        file_frame.rowconfigure(1, weight=1)

        # Drag and drop variables
        self.drag_data = {"item": None, "index": None}

        # Bind drag and drop events
        self.file_listbox.bind("<Button-1>", self.on_drag_start)
        self.file_listbox.bind("<B1-Motion>", self.on_drag_motion)
        self.file_listbox.bind("<ButtonRelease-1>", self.on_drag_drop)

        self.file_listbox.bind("<<ListboxSelect>>", self.update_sheet_list)

        # 시트 프레임
        sheet_frame = ttk.LabelFrame(main_frame, text="시트 목록")
        sheet_frame.grid(row=0, column=3, rowspan=2, padx=5, pady=5, sticky=(tk.W, tk.E, tk.N, tk.S))
        main_frame.columnconfigure(3, weight=1)

        # 시트 목록 리스트박스
        self.sheet_listbox = tk.Listbox(sheet_frame, selectmode=tk.EXTENDED)
        self.sheet_listbox.grid(row=0, column=0, padx=5, pady=5, sticky=(tk.W, tk.E, tk.N, tk.S))
        sheet_frame.columnconfigure(0, weight=1)
        sheet_frame.rowconfigure(0, weight=1)

        # 미리보기 버튼
        self.preview_button = ttk.Button(sheet_frame, text="미리보기", command=self.preview_sheet)
        self.preview_button.grid(row=1, column=0, padx=5, pady=5, sticky=tk.E)

        # 병합 버튼
        self.merge_button = ttk.Button(main_frame, text="선택된 시트 병합", command=self.merge_sheets)
        self.merge_button.grid(row=2, column=0, columnspan=4, padx=5, pady=5, sticky=(tk.W, tk.E))

    def select_files():
        files = filedialog.askopenfilenames(
            title="병합할 엑셀 파일을 선택하세요",
            filetypes=(("Excel files", "*.xlsx *.xls"), ("All files", "*.*" ))
        )
        if files:
            for file_path in files:
                if file_path in self.file_listbox.get(0, tk.END): # Avoid adding duplicates
                    continue

                password = None
                is_encrypted = False

                # Try to open without password first
                try:
                    if file_path.endswith('.xlsx'):
                        openpyxl.load_workbook(file_path, read_only=True)
                    elif file_path.endswith('.xls'):
                        xlrd.open_workbook(file_path)
                    # If successful, it's not encrypted or already handled
                    self.file_listbox.insert(tk.END, file_path)
                except InvalidFileException: # openpyxl encryption error
                    is_encrypted = True
                except xlrd.biffh.XLRDError as e: # xlrd encryption error
                    if "Bad password" in str(e) or "File is encrypted" in str(e):
                        is_encrypted = True
                    else:
                        print(f"Error opening {file_path}: {e}")
                        continue
                except Exception as e:
                    print(f"Error opening {file_path}: {e}")
                    continue

                if is_encrypted:
                    password = self.ask_for_password(file_path)
                    if password:
                        try:
                            if file_path.endswith('.xlsx'):
                                openpyxl.load_workbook(file_path, read_only=True, password=password)
                            elif file_path.endswith('.xls'):
                                xlrd.open_workbook(file_path, password=password)
                            self.file_passwords[file_path] = password
                            self.file_listbox.insert(tk.END, file_path)
                        except (InvalidFileException, xlrd.biffh.XLRDError) as e:
                            messagebox.showerror("암호 오류", f"'{file_path}' 파일의 암호가 올바르지 않습니다.")
                        except Exception as e:
                            messagebox.showerror("파일 오류", f"'{file_path}' 파일을 여는 중 오류가 발생했습니다: {e}")
                    else:
                        messagebox.showinfo("취소", f"'{file_path}' 파일 추가를 취소했습니다.")

    def ask_for_password(self, file_path):
        password = simpledialog.askstring("암호 입력", f"'{file_path}' 파일은 암호화되어 있습니다. 암호를 입력하세요:", show='*')
        return password

    def exclude_files(self):
        selected_indices = self.file_listbox.curselection()
        for index in reversed(selected_indices):
            file_path = self.file_listbox.get(index)
            if file_path in self.file_passwords:
                del self.file_passwords[file_path]
            if file_path in self.selected_sheets:
                del self.selected_sheets[file_path]
            self.file_listbox.delete(index)

    def move_up(self):
        selected_indices = self.file_listbox.curselection()
        for index in selected_indices:
            if index > 0:
                item = self.file_listbox.get(index)
                self.file_listbox.delete(index)
                self.file_listbox.insert(index - 1, item)
                self.file_listbox.selection_set(index - 1)

    def move_down(self):
        selected_indices = self.file_listbox.curselection()
        for index in reversed(selected_indices):
            if index < self.file_listbox.size() - 1:
                item = self.file_listbox.get(index)
                self.file_listbox.delete(index)
                self.file_listbox.insert(index + 1, item)
                self.file_listbox.selection_set(index + 1)

    def on_drag_start(self, event):
        # Get the index of the item clicked
        index = self.file_listbox.nearest(event.y)
        if index != '':
            self.drag_data["item"] = self.file_listbox.get(index)
            self.drag_data["original_index"] = index # Store original index
            self.file_listbox.selection_clear(0, tk.END)
            self.file_listbox.selection_set(index) # Select the dragged item

    def on_drag_motion(self, event):
        if self.drag_data["item"] is not None:
            current_index = self.file_listbox.nearest(event.y)
            if current_index != '':
                self.file_listbox.selection_clear(0, tk.END)
                self.file_listbox.selection_set(current_index) # Highlight the potential drop target

    def on_drag_drop(self, event):
        if self.drag_data["item"] is not None:
            drop_index = self.file_listbox.nearest(event.y)
            original_index = self.drag_data["original_index"]
            dragged_item = self.drag_data["item"]

            if drop_index != '' and original_index != drop_index:
                # Remove the item from its original position
                self.file_listbox.delete(original_index)
                # Insert the item at the new position
                self.file_listbox.insert(drop_index, dragged_item)
                self.file_listbox.selection_clear(0, tk.END)
                self.file_listbox.selection_set(drop_index) # Select the dropped item

        self.drag_data = {"item": None, "original_index": None} # Reset drag data


    def update_sheet_list(self, event):
        selected_indices = self.file_listbox.curselection()
        if not selected_indices:
            self.sheet_listbox.delete(0, tk.END)
            return

        self.sheet_listbox.delete(0, tk.END)

        selected_file = self.file_listbox.get(selected_indices[0])
        password = self.file_passwords.get(selected_file) # Get stored password

        try:
            if selected_file.endswith('.xlsx'):
                workbook = openpyxl.load_workbook(selected_file, read_only=True, password=password)
                sheet_names = workbook.sheetnames
            elif selected_file.endswith('.xls'):
                workbook = xlrd.open_workbook(selected_file, password=password)
                sheet_names = workbook.sheet_names()
            else:
                return

            # Clear previous selections for this file
            self.selected_sheets[selected_file] = []

            for i, sheet_name in enumerate(sheet_names):
                self.sheet_listbox.insert(tk.END, sheet_name)
                self.sheet_listbox.selection_set(i) # Select all sheets by default
                self.selected_sheets[selected_file].append(sheet_name)

        except (InvalidFileException, xlrd.biffh.XLRDError) as e:
            messagebox.showerror("암호 오류", f"'{selected_file}' 파일의 암호가 올바르지 않거나 파일이 손상되었습니다.")
        except Exception as e:
            print(f"Error reading file: {e}")

    def on_sheet_select(self, event):
        selected_file_indices = self.file_listbox.curselection()
        if not selected_file_indices:
            return

        selected_file = self.file_listbox.get(selected_file_indices[0])
        current_selected_sheets = [self.sheet_listbox.get(i) for i in self.sheet_listbox.curselection()]
        self.selected_sheets[selected_file] = current_selected_sheets

    def preview_sheet(self):
        selected_file_indices = self.file_listbox.curselection()
        selected_sheet_indices = self.sheet_listbox.curselection()

        if not selected_file_indices or not selected_sheet_indices:
            messagebox.showwarning("선택 오류", "미리보기할 파일과 시트를 선택해주세요.")
            return

        selected_file_path = self.file_listbox.get(selected_file_indices[0])
        selected_sheet_name = self.sheet_listbox.get(selected_sheet_indices[0])
        password = self.file_passwords.get(selected_file_path)

        try:
            # Use pandas to read the Excel sheet into a DataFrame
            # This handles both .xlsx and .xls and passwords
            if selected_file_path.endswith('.xlsx'):
                df = pd.read_excel(selected_file_path, sheet_name=selected_sheet_name, engine='openpyxl', password=password)
            elif selected_file_path.endswith('.xls'):
                df = pd.read_excel(selected_file_path, sheet_name=selected_sheet_name, engine='xlrd', password=password)
            else:
                messagebox.showerror("파일 형식 오류", "지원되지 않는 파일 형식입니다.")
                return

            self.show_preview_window(selected_file_path, selected_sheet_name, df)

        except Exception as e:
            messagebox.showerror("미리보기 오류", f"시트를 미리 볼 수 없습니다: {e}")

    def show_preview_window(self, file_path, sheet_name, dataframe):
        preview_window = tk.Toplevel(self)
        preview_window.title(f"미리보기: {sheet_name} ({file_path.split('/')[-1]})")
        preview_window.geometry("800x600")

        # Create a Treeview widget
        tree = ttk.Treeview(preview_window)
        tree.pack(expand=True, fill='both')

        # Define columns
        tree["columns"] = list(dataframe.columns)
        tree["show"] = "headings"

        # Add column headings
        for col in dataframe.columns:
            tree.heading(col, text=str(col))
            tree.column(col, width=100) # Default width

        # Add data to the treeview
        for index, row in dataframe.iterrows():
            tree.insert("", "end", values=list(row))

        # Add scrollbars
        vsb = ttk.Scrollbar(tree, orient="vertical", command=tree.yview)
        hsb = ttk.Scrollbar(tree, orient="horizontal", command=tree.xview)
        tree.configure(yscrollcommand=vsb.set, xscrollcommand=hsb.set)
        vsb.pack(side='right', fill='y')
        hsb.pack(side='bottom', fill='x')

    def merge_sheets(self):
        messagebox.showinfo("병합", "시트 병합 기능은 아직 구현되지 않았습니다.")

if __name__ == "__main__":
    app = ExcelMergerApp()
    app.mainloop()
