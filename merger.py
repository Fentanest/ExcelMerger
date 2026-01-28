import openpyxl
import xlrd
import re
import os
from PySide6.QtWidgets import QApplication

class Merger:
    def __init__(self, main_window):
        self.main_window = main_window

    def merge_as_sheets(self, sheets_to_merge, save_path):
        output_workbook = openpyxl.Workbook()
        output_workbook.remove(output_workbook.active) # Remove default sheet

        if self.main_window.options['only_value_copy']:
            self.main_window.txtLogOutput.append("수식을 값으로 변환하며 병합을 시작합니다...")

        total_sheets = len(sheets_to_merge)
        for i, item in enumerate(sheets_to_merge):
            file_name, sheet_name = item.split('/', 1)
            file_path = self.main_window.file_info.get(file_name, {}).get('processed_path')
            
            if not file_path:
                self.main_window.txtLogOutput.append(f"파일을 찾을 수 없습니다: {file_name}")
                continue

            if self.main_window.debug_mode and file_name in self.main_window.file_passwords:
                self.main_window.txtLogOutput.append(f"DEBUG: {file_name}에 기억된 비밀번호로 복호화된 임시파일 열기 시도...")

            try:
                source_workbook = self.main_window.file_handler._open_workbook(file_path, file_name, data_only=self.main_window.options['only_value_copy']) # Pass data_only flag
                if not source_workbook:
                    continue

                if file_path.lower().endswith(('.xlsx', '.xlsm', '.xlsb', '.csv')):
                    source_sheet = source_workbook[sheet_name]
                else: # .xls
                    source_sheet = source_workbook.sheet_by_name(sheet_name)
                sheet_name_rule = self.main_window.options.get('sheet_name_rule', 'OriginalBoth')
                if sheet_name_rule == 'OriginalSheet':
                    new_sheet_name = sheet_name
                elif sheet_name_rule == 'OriginalFileName':
                    new_sheet_name = os.path.splitext(file_name)[0]
                else: # OriginalBoth
                    new_sheet_name = f"{os.path.splitext(file_name)[0]}_{sheet_name}"

                # Sanitize and truncate
                if len(new_sheet_name) > 31:
                    self.main_window.txtLogOutput.append(f"시트 이름이 31자를 초과하여 일부 잘립니다: {new_sheet_name}")
                    new_sheet_name = new_sheet_name[:31]
                new_sheet_name = re.sub(r'[\\\\[\\\\]\\*\\:\\?/]', '_', new_sheet_name)

                # Handle duplicates for all sheet name rules
                original_new_sheet_name = new_sheet_name
                counter = 2
                while new_sheet_name in output_workbook.sheetnames:
                    suffix = f" ({counter})"
                    truncated_len = 31 - len(suffix)
                    new_sheet_name = f"{original_new_sheet_name[:truncated_len]}{suffix}"
                    counter += 1
                
                output_sheet = output_workbook.create_sheet(title=new_sheet_name)
                self.copy_sheet_data(source_sheet, output_sheet, file_name=file_name)

            except Exception as e:
                self.main_window.txtLogOutput.append(f"시트 복사 오류 {item}: {e}")

            self.main_window.progressBar.setValue(int((i + 1) / total_sheets * 100))
            QApplication.processEvents()

        self.perform_sheet_trim(output_workbook)

        output_workbook.save(save_path)

    def _merge_by_axis(self, sheets_to_merge, save_path, axis):
        output_workbook = openpyxl.Workbook()
        output_sheet = output_workbook.active
        output_sheet.title = "Merged_Sheet"

        if self.main_window.options['only_value_copy']:
            self.main_window.txtLogOutput.append("수식을 값으로 변환하며 병합을 시작합니다...")

        total_sheets = len(sheets_to_merge)
        last_pos = 0
        for i, item in enumerate(sheets_to_merge):
            file_name, sheet_name = item.split('/', 1)
            self.main_window.lblCurrentFile.setText(f'{item} 병합 중...')

            file_path = self.main_window.file_info.get(file_name, {}).get('processed_path')

            if not file_path:
                self.main_window.txtLogOutput.append(f"파일을 찾을 수 없습니다: {file_name}")
                continue

            if self.main_window.debug_mode and file_name in self.main_window.file_passwords:
                self.main_window.txtLogOutput.append(f"DEBUG: {file_name}에 기억된 비밀번호로 복호화된 임시파일 열기 시도...")

            try:
                source_workbook = self.main_window.file_handler._open_workbook(file_path, file_name, data_only=self.main_window.options['only_value_copy'])
                if not source_workbook:
                    continue

                if file_path.lower().endswith(('.xlsx', '.xlsm', '.xlsb', '.csv')):
                    source_sheet = source_workbook[sheet_name]
                else: # .xls
                    source_sheet = source_workbook.sheet_by_name(sheet_name)
                
                if axis == 'horizontal':
                    self.copy_sheet_data(source_sheet, output_sheet, start_col=last_pos + 1, file_name=file_name)
                    last_pos = output_sheet.max_column
                else: # vertical
                    self.copy_sheet_data(source_sheet, output_sheet, start_row=last_pos + 1, file_name=file_name)
                    last_pos = output_sheet.max_row

            except Exception as e:
                self.main_window.txtLogOutput.append(f"시트 병합 오류 {item}: {e}")

            self.main_window.progressBar.setValue(int((i + 1) / total_sheets * 100))
            QApplication.processEvents()

        self.perform_sheet_trim(output_workbook)

        output_workbook.save(save_path)

    def merge_horizontally(self, sheets_to_merge, save_path):
        self._merge_by_axis(sheets_to_merge, save_path, 'horizontal')

    def merge_vertically(self, sheets_to_merge, save_path):
        self._merge_by_axis(sheets_to_merge, save_path, 'vertical')

    def copy_sheet_data(self, source_sheet, output_sheet, start_row=1, start_col=1, file_name=""):
        if isinstance(source_sheet, openpyxl.worksheet.worksheet.Worksheet):
            for row in source_sheet.iter_rows():
                for cell in row:
                    new_cell = output_sheet.cell(row=cell.row + start_row - 1, column=cell.column + start_col - 1)
                    new_cell.value = cell.value
                    if cell.has_style:
                        new_cell.font = cell.font.copy()
                        new_cell.border = cell.border.copy()
                        new_cell.fill = cell.fill.copy()
                        new_cell.number_format = cell.number_format
                        new_cell.protection = cell.protection.copy()
                        new_cell.alignment = cell.alignment.copy()
        elif isinstance(source_sheet, xlrd.sheet.Sheet):
            self.main_window.txtLogOutput.append(f".xls 파일({file_name}/{source_sheet.name})의 서식은 일부만 지원됩니다.")
            for row_idx in range(source_sheet.nrows):
                for col_idx in range(source_sheet.ncols):
                    cell_value = source_sheet.cell_value(row_idx, col_idx)
                    output_sheet.cell(row=row_idx + start_row, column=col_idx + start_col).value = cell_value

    def perform_sheet_trim(self, workbook):
        sheet_trim_value = self.main_window.options.get('sheet_trim_value', 0)
        if sheet_trim_value <= 0:
            return

        trim_rows = self.main_window.options.get('sheet_trim_rows', False)
        trim_cols = self.main_window.options.get('sheet_trim_cols', False)
        if not trim_rows and not trim_cols:
            return

        self.main_window.txtLogOutput.append("시트 정리(SheetTrim) 기능 수행 중...")
        
        for worksheet in workbook.worksheets:
            if trim_rows:
                empty_row_indices = []
                for i in range(1, worksheet.max_row + 1):
                    if all(c.value is None or str(c.value).strip() == '' for c in worksheet[i]):
                        empty_row_indices.append(i)
                
                if empty_row_indices:
                    blocks = []
                    start_of_block = empty_row_indices[0]
                    for i in range(1, len(empty_row_indices)):
                        if empty_row_indices[i] != empty_row_indices[i-1] + 1:
                            block_len = empty_row_indices[i-1] - start_of_block + 1
                            if block_len >= sheet_trim_value:
                                blocks.append((start_of_block, block_len))
                            start_of_block = empty_row_indices[i]
                    block_len = empty_row_indices[-1] - start_of_block + 1
                    if block_len >= sheet_trim_value:
                        blocks.append((start_of_block, block_len))

                    for start, count in reversed(blocks):
                        worksheet.delete_rows(start, count)

            if trim_cols:
                empty_col_indices = []
                for i, col in enumerate(worksheet.iter_cols(), 1):
                    if all(cell.value is None or str(cell.value).strip() == "" for cell in col):
                        empty_col_indices.append(i)

                if empty_col_indices:
                    blocks = []
                    start_of_block = empty_col_indices[0]
                    for i in range(1, len(empty_col_indices)):
                        if empty_col_indices[i] != empty_col_indices[i-1] + 1:
                            block_len = empty_col_indices[i-1] - start_of_block + 1
                            if block_len >= sheet_trim_value:
                                blocks.append((start_of_block, block_len))
                            start_of_block = empty_col_indices[i]
                    block_len = empty_col_indices[-1] - start_of_block + 1
                    if block_len >= sheet_trim_value:
                        blocks.append((start_of_block, block_len))

                    for start, count in reversed(blocks):
                        worksheet.delete_cols(start, count)

        self.main_window.txtLogOutput.append("시트 정리(SheetTrim) 완료.")
