import csv
import getpass
import io
import os
import sys
import tempfile

try:
    import msoffcrypto
except ImportError:
    msoffcrypto = None

from excelmerger.file_registry import build_password_cache_key

try:
    import openpyxl
except ImportError:
    openpyxl = None

try:
    import xlrd
except ImportError:
    xlrd = None

try:
    from pyxlsb import open_workbook as open_xlsb
except ImportError:
    open_xlsb = None

class FileHandler:
    def __init__(self, main_window):
        self.main_window = main_window
        if self.main_window is not None:
            self.main_window.temp_files = getattr(self.main_window, "temp_files", [])
            self.main_window.file_passwords = getattr(self.main_window, "file_passwords", {})
            self.main_window.global_password = getattr(self.main_window, "global_password", "")
            self.main_window.use_global_password = getattr(self.main_window, "use_global_password", False)
            self.main_window.stop_asking_for_passwords = getattr(
                self.main_window,
                "stop_asking_for_passwords",
                False,
            )

    def _log(self, message):
        log_output = getattr(self.main_window, "txtLogOutput", None)
        if log_output is not None and hasattr(log_output, "append"):
            log_output.append(message)

    def _debug_enabled(self):
        return bool(getattr(self.main_window, "debug_mode", False))

    def _password_prompt_result(self, basename):
        provider = getattr(self.main_window, "password_provider", None)
        if callable(provider):
            provided = provider(basename)
            if isinstance(provided, dict):
                return {
                    "password": provided.get("password", ""),
                    "remember_globally": bool(provided.get("remember_globally", False)),
                    "stopped": bool(provided.get("stopped", False)),
                    "cancelled": bool(provided.get("cancelled", False)),
                }
            if provided is None:
                return {"password": "", "remember_globally": False, "stopped": False, "cancelled": True}
            return {
                "password": str(provided),
                "remember_globally": False,
                "stopped": False,
                "cancelled": False,
            }

        try:
            from excelmerger.ui.dialogs import PasswordDialog
        except Exception:
            PasswordDialog = None

        if PasswordDialog is not None and self.main_window is not None and hasattr(self.main_window, "windowIcon"):
            dialog = PasswordDialog(basename, self.main_window)
            dialog.setWindowIcon(self.main_window.windowIcon())
            result = dialog.exec()
            if dialog.stopped:
                return {"password": "", "remember_globally": False, "stopped": True, "cancelled": False}
            if not result:
                return {"password": "", "remember_globally": False, "stopped": False, "cancelled": True}
            return {
                "password": dialog.lineEditKeepPassword.text(),
                "remember_globally": dialog.chkKeepPassword.isChecked(),
                "stopped": False,
                "cancelled": False,
            }

        if not sys.stdin.isatty():
            self._log(f"비대화형 환경에서는 {basename}의 비밀번호를 입력받을 수 없습니다.")
            return {"password": "", "remember_globally": False, "stopped": False, "cancelled": True}

        try:
            password = getpass.getpass(f"{basename} 비밀번호: ")
        except (EOFError, KeyboardInterrupt):
            self._log("터미널 비밀번호 입력이 취소되었습니다.")
            return {"password": "", "remember_globally": False, "stopped": False, "cancelled": True}

        if not password:
            return {"password": "", "remember_globally": False, "stopped": False, "cancelled": True}

        return {"password": password, "remember_globally": False, "stopped": False, "cancelled": False}

    def _open_workbook(self, file_path, file_name, data_only=False):
        try:
            lower_path = file_path.lower()
            if lower_path.endswith('.xlsx'):
                if openpyxl is None:
                    raise RuntimeError("openpyxl이 설치되어 있지 않습니다.")
                return openpyxl.load_workbook(file_path, read_only=False, data_only=data_only)
            elif lower_path.endswith('.xlsm'):
                if openpyxl is None:
                    raise RuntimeError("openpyxl이 설치되어 있지 않습니다.")
                return openpyxl.load_workbook(file_path, read_only=False, keep_vba=True, data_only=data_only)
            elif lower_path.endswith('.xls'):
                if xlrd is None:
                    raise RuntimeError("xlrd가 설치되어 있지 않습니다.")
                return xlrd.open_workbook(file_path, formatting_info=True)
            elif lower_path.endswith('.xlsb'):
                if open_xlsb is None or openpyxl is None:
                    raise RuntimeError("pyxlsb 또는 openpyxl이 설치되어 있지 않습니다.")
                self.main_window.txtLogOutput.append(f"표준 병합을 위해 .xlsb 파일을 변환 중: {file_name}")
                with open_xlsb(file_path) as wb_xlsb:
                    wb_xlsx = openpyxl.Workbook()
                    wb_xlsx.remove(wb_xlsx.active)
                    for sheet_name in wb_xlsb.sheets:
                        ws_xlsx = wb_xlsx.create_sheet(sheet_name)
                        with wb_xlsb.get_sheet(sheet_name) as sheet_xlsb:
                            for row in sheet_xlsb.rows():
                                ws_xlsx.append([c.v for c in row])
                    return wb_xlsx
            elif lower_path.endswith('.csv'):
                if openpyxl is None:
                    raise RuntimeError("openpyxl이 설치되어 있지 않습니다.")
                self._log(f"표준 병합을 위해 .csv 파일을 변환 중: {file_name}")
                wb = openpyxl.Workbook()
                ws = wb.active
                ws.title = os.path.splitext(file_name)[0]
                try:
                    with open(file_path, 'r', newline='', encoding='utf-8') as csvfile:
                        reader = csv.reader(csvfile)
                        for row in reader:
                            ws.append(row)
                except UnicodeDecodeError:
                    with open(file_path, 'r', newline='', encoding='cp949') as csvfile:
                        reader = csv.reader(csvfile)
                        for row in reader:
                            ws.append(row)
                return wb
        except Exception as e:
            self._log(f"파일 열기 오류 {file_name}: {e}")
            return None
        return None

    def convert_to_xlsx(self, file_path):
        file_name = os.path.basename(file_path)
        lower_file_path = file_path.lower()

        if sys.platform == 'win32' and self.main_window.merger_win32.win32 and lower_file_path.endswith(('.xls', '.xlsb')):
            return self.main_window.merger_win32.convert_to_xlsx_win32(file_path)

        merger_poi = getattr(self.main_window, 'merger_poi', None)
        if (
            merger_poi is not None
            and merger_poi.is_available()
            and lower_file_path.endswith(('.xls', '.xlsb', '.xlsm', '.csv'))
        ):
            try:
                self._log(f"POI로 .xlsx 변환 중: {file_name}")
                xlsx_path = merger_poi.convert_to_xlsx(file_path)
                if xlsx_path:
                    self.main_window.temp_files.append(xlsx_path)
                    return xlsx_path
            except Exception as exc:
                self._log(f"POI 변환 실패, openpyxl 폴백 시도: {exc}")

        if lower_file_path.endswith('.csv'):
            try:
                if openpyxl is None:
                    raise RuntimeError("CSV 변환에 필요한 openpyxl이 설치되어 있지 않습니다.")
                self._log(f".csv 파일을 .xlsx로 변환 중: {file_name}")
                wb = openpyxl.Workbook()
                ws = wb.active
                
                try:
                    with open(file_path, 'r', newline='', encoding='utf-8') as csvfile:
                        reader = csv.reader(csvfile)
                        for row in reader:
                            ws.append(row)
                except UnicodeDecodeError:
                    with open(file_path, 'r', newline='', encoding='cp949') as csvfile: # Fallback for Korean
                        reader = csv.reader(csvfile)
                        for row in reader:
                            ws.append(row)

                fd, xlsx_path = tempfile.mkstemp(suffix='.xlsx', prefix='csv_converted_')
                os.close(fd)
                wb.save(xlsx_path)
                self.main_window.temp_files.append(xlsx_path)
                return xlsx_path
            except Exception as e:
                self._log(f".csv to .xlsx 변환 오류: {e}")
                return None

        try:
            fd, xlsx_path = tempfile.mkstemp(suffix='.xlsx', prefix='converted_')
            os.close(fd)
            
            if lower_file_path.endswith('.xlsb'):
                if open_xlsb is None or openpyxl is None:
                    raise RuntimeError("pyxlsb 또는 openpyxl이 설치되어 있지 않습니다.")
                self._log(f".xlsb 파일을 .xlsx로 변환 중: {file_name}")
                with open_xlsb(file_path) as wb_xlsb:
                    wb_xlsx = openpyxl.Workbook()
                    wb_xlsx.remove(wb_xlsx.active)
                    for sheet_name in wb_xlsb.sheets:
                        ws_xlsx = wb_xlsx.create_sheet(sheet_name)
                        with wb_xlsb.get_sheet(sheet_name) as sheet_xlsb:
                            for row in sheet_xlsb.rows():
                                ws_xlsx.append([c.v for c in row])
                    wb_xlsx.save(xlsx_path)
                self.main_window.temp_files.append(xlsx_path)
                return xlsx_path

            elif lower_file_path.endswith('.xlsm'):
                if openpyxl is None:
                    raise RuntimeError("openpyxl이 설치되어 있지 않습니다.")
                self._log(f".xlsm 파일을 .xlsx로 변환 중 (VBA 제외): {file_name}")
                wb = openpyxl.load_workbook(file_path, data_only=True)
                wb.save(xlsx_path)
                self.main_window.temp_files.append(xlsx_path)
                return xlsx_path

            elif lower_file_path.endswith('.xls'):
                if xlrd is None or openpyxl is None:
                    raise RuntimeError("xlrd 또는 openpyxl이 설치되어 있지 않습니다.")
                self._log(f".xls 파일을 .xlsx로 변환 중: {file_name}")
                wb_xls = xlrd.open_workbook(file_path)
                wb_xlsx = openpyxl.Workbook()
                wb_xlsx.remove(wb_xlsx.active)
                for sheet_xls in wb_xls.sheets():
                    ws_xlsx = wb_xlsx.create_sheet(sheet_xls.name)
                    for row in range(sheet_xls.nrows):
                        ws_xlsx.append(sheet_xls.row_values(row))
                wb_xlsx.save(xlsx_path)
                self.main_window.temp_files.append(xlsx_path)
                return xlsx_path

        except Exception as e:
            self._log(f"파일 변환 오류 ({file_name}): {e}")
            return None
        
        return file_path

    def get_sheet_names(self, file_path):
        file_name = os.path.basename(file_path)
        processed_file_path = file_path
        file_ext = os.path.splitext(file_name)[1].lower()

        if file_ext in ['.xlsx', '.xls', '.xlsm', '.xlsb']:
            is_encrypted = False
            if msoffcrypto is not None:
                try:
                    with open(processed_file_path, 'rb') as f:
                        office_file = msoffcrypto.OfficeFile(f)
                        if office_file.is_encrypted():
                            is_encrypted = True
                except Exception:
                    pass # Not a valid office file

            if is_encrypted:
                decrypted_temp_path = self.handle_encrypted_file(processed_file_path)
                if decrypted_temp_path:
                    processed_file_path = decrypted_temp_path
                else:
                    self._log(f"파일을 열 수 없습니다 (암호화 문제 또는 사용자 취소): {file_name}")
                    return None, None

        self._log(f"시트 목록 읽기: {file_name}")
        try:
            lower_path = processed_file_path.lower()
            if lower_path.endswith('.csv'):
                return [os.path.splitext(file_name)[0]], processed_file_path

            if lower_path.endswith('.xlsb'):
                converted_path = self.convert_to_xlsx(processed_file_path)
                if not converted_path:
                    return None, None
                processed_file_path = converted_path
                lower_path = processed_file_path.lower()

            if hasattr(self.main_window, 'merger_poi') and self.main_window.merger_poi.is_available():
                try:
                    return self.main_window.merger_poi.get_sheet_names(processed_file_path), processed_file_path
                except Exception as exc:
                    self._log(f"JPype 시트 목록 읽기 실패, 표준 모드로 전환합니다: {exc}")

            if lower_path.endswith('.xlsm'):
                if openpyxl is None:
                    raise RuntimeError("openpyxl이 설치되어 있지 않습니다.")
                wb = openpyxl.load_workbook(processed_file_path, read_only=False, keep_vba=True, data_only=True)
                sheet_names = wb.sheetnames
                wb.close()
                return sheet_names, processed_file_path
            elif lower_path.endswith('.xlsx'):
                if openpyxl is None:
                    raise RuntimeError("openpyxl이 설치되어 있지 않습니다.")
                wb = openpyxl.load_workbook(processed_file_path, read_only=True, data_only=True)
                sheet_names = wb.sheetnames
                wb.close()
                return sheet_names, processed_file_path
            elif lower_path.endswith('.xls'):
                if xlrd is None:
                    raise RuntimeError("xlrd가 설치되어 있지 않습니다.")
                wb = xlrd.open_workbook(processed_file_path, on_demand=True)
                return wb.sheet_names(), processed_file_path
            elif lower_path.endswith('.xlsb'):
                if open_xlsb is None:
                    raise RuntimeError("pyxlsb가 설치되어 있지 않습니다.")
                with open_xlsb(processed_file_path) as wb:
                    return wb.sheets, processed_file_path
            else:
                self._log(f"지원하지 않는 파일 형식입니다: {file_name}")
                return None, None
                
        except Exception as e:
            self._log(f"시트 이름 가져오기 오류 ({file_name}): {e}")
            return None, None

    def handle_encrypted_file(self, file_path):
        if msoffcrypto is None:
            self._log("암호화된 Office 파일을 처리하려면 msoffcrypto-tool이 필요합니다.")
            return None

        if self.main_window.stop_asking_for_passwords:
            self._log(f"비밀번호 입력을 중단하여 파일을 건너뜁니다: {os.path.basename(file_path)}")
            if self._debug_enabled():
                self._log("DEBUG: handle_encrypted_file returning None (stopped asking).")
            return None

        self._log(f"암호화된 파일 감지: {os.path.basename(file_path)}")
        password = None
        decrypted_file_buffer = None
        basename = os.path.basename(file_path)
        cache_key = build_password_cache_key(file_path)
        temp_decrypted_path = None

        if cache_key in self.main_window.file_passwords:
            try:
                self._log(f'{basename}에 대해 기억된 비밀번호로 열기 시도...')
                decrypted_file_buffer = io.BytesIO()
                with open(file_path, 'rb') as f:
                    office_file = msoffcrypto.OfficeFile(f)
                    office_file.load_key(password=self.main_window.file_passwords[cache_key])
                    office_file.decrypt(decrypted_file_buffer)
                password = self.main_window.file_passwords[cache_key]
                self._log("기억된 비밀번호로 열기 성공.")
            except Exception:
                self._log("기억된 비밀번호 실패.")
                decrypted_file_buffer = None

        if not password and self.main_window.use_global_password and self.main_window.global_password:
            try:
                self._log("전역 비밀번호로 열기 시도...")
                decrypted_file_buffer = io.BytesIO()
                with open(file_path, 'rb') as f:
                    office_file = msoffcrypto.OfficeFile(f)
                    office_file.load_key(password=self.main_window.global_password)
                    office_file.decrypt(decrypted_file_buffer)
                password = self.main_window.global_password
                self.main_window.file_passwords[cache_key] = self.main_window.global_password
                self._log("전역 비밀번호로 열기 성공.")
            except Exception as e:
                self._log(f"전역 비밀번호 실패: {e}")
                decrypted_file_buffer = None

        if not password:
            if self._debug_enabled():
                self._log(f"DEBUG: Requesting password for {basename}.")

            prompt_result = self._password_prompt_result(basename)
            if prompt_result["stopped"]:
                self.main_window.stop_asking_for_passwords = True
                self._log("사용자가 중단하여 이후 암호 입력을 건너뜁니다.")
                if self._debug_enabled():
                    self._log("DEBUG: handle_encrypted_file returning None (dialog stopped).")
                return None

            if not prompt_result["cancelled"]:
                user_password = prompt_result["password"]
                if user_password:
                    try:
                        self._log("사용자 입력 비밀번호로 열기 시도...")
                        decrypted_file_buffer = io.BytesIO()
                        with open(file_path, 'rb') as f:
                            office_file = msoffcrypto.OfficeFile(f)
                            office_file.load_key(password=user_password)
                            office_file.decrypt(decrypted_file_buffer)
                        password = user_password
                        self._log("사용자 입력 비밀번호로 열기 성공.")
                        self.main_window.file_passwords[cache_key] = user_password
                        if prompt_result["remember_globally"]:
                            self.main_window.global_password = user_password
                            self.main_window.use_global_password = True
                    except Exception as e:
                        self._log(f"사용자 입력 비밀번호 실패: {e}")
                        if self._debug_enabled():
                            self._log("DEBUG: handle_encrypted_file returning None (user password failed).")
                        return None
                else:
                    self._log("비밀번호가 입력되지 않아 파일을 건너뜁니다.")
                    if self._debug_enabled():
                        self._log("DEBUG: handle_encrypted_file returning None (empty user password).")
                    return None
            else:
                self._log("사용자가 취소하여 파일을 건너뜁니다.")
                if self._debug_enabled():
                    self._log("DEBUG: handle_encrypted_file returning None (dialog cancelled).")
                return None

        if password and decrypted_file_buffer:
            fd, temp_decrypted_path = tempfile.mkstemp(suffix=os.path.splitext(file_path)[1], prefix='decrypted_')
            os.close(fd)
            with open(temp_decrypted_path, 'wb') as tmp_f:
                tmp_f.write(decrypted_file_buffer.getbuffer())
            self.main_window.temp_files.append(temp_decrypted_path)
            if self._debug_enabled():
                self._log(f"DEBUG: handle_encrypted_file returning decrypted path: {temp_decrypted_path}")
            return temp_decrypted_path
        
        if self._debug_enabled():
            self._log("DEBUG: handle_encrypted_file returning None (no password worked).")
        return None
