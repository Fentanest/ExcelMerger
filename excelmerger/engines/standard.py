import os
from copy import copy

from excelmerger.engines.utils import build_output_sheet_name
from excelmerger.file_registry import source_file_name

try:
    import openpyxl
except ImportError:
    openpyxl = None

try:
    import xlrd
except ImportError:
    xlrd = None


class Merger:
    """openpyxl 기반 표준 병합 엔진. POI/Excel/LibreOffice가 모두 사용 불가할 때 최후 폴백."""

    def __init__(self, main_window=None, log_callback=None, progress_callback=None, status_callback=None):
        self.main_window = main_window
        self.log_callback = log_callback
        self.progress_callback = progress_callback
        self.status_callback = status_callback

    def _log(self, message):
        if self.log_callback:
            self.log_callback(message)
        elif self.main_window and hasattr(self.main_window, "txtLogOutput"):
            self.main_window.txtLogOutput.append(message)

    def _progress(self, percent):
        if self.progress_callback:
            self.progress_callback(percent)
        elif self.main_window and hasattr(self.main_window, "progressBar"):
            self.main_window.progressBar.setValue(percent)

    def _status(self, text):
        if self.status_callback:
            self.status_callback(text)
        elif self.main_window and hasattr(self.main_window, "lblCurrentFile"):
            self.main_window.lblCurrentFile.setText(text)

    def is_available(self):
        return openpyxl is not None

    def _require_openpyxl(self):
        if openpyxl is None:
            raise RuntimeError("표준 병합 엔진에 필요한 openpyxl이 설치되어 있지 않습니다.")

    def merge_as_sheets(self, sheets_to_merge, save_path):
        self._require_openpyxl()

        output_workbook = openpyxl.Workbook()
        output_workbook.remove(output_workbook.active)

        if self.main_window.options['only_value_copy']:
            self._log("수식을 값으로 변환하며 병합을 시작합니다...")

        total_sheets = len(sheets_to_merge)
        sheet_errors = []
        merged_count = 0
        for i, item in enumerate(sheets_to_merge):
            file_name, sheet_name = item.split('/', 1)
            file_path = self.main_window.file_info.get(file_name, {}).get('processed_path')
            info = self.main_window.file_info.get(file_name, {})

            if not file_path:
                self._log(f"파일을 찾을 수 없습니다: {file_name}")
                sheet_errors.append(item)
                continue

            try:
                source_workbook = self.main_window.file_handler._open_workbook(
                    file_path, file_name, data_only=self.main_window.options['only_value_copy']
                )
                if not source_workbook:
                    continue

                if file_path.lower().endswith(('.xlsx', '.xlsm', '.xlsb', '.csv')):
                    source_sheet = source_workbook[sheet_name]
                else:
                    source_sheet = source_workbook.sheet_by_name(sheet_name)

                new_sheet_name = build_output_sheet_name(
                    source_file_name(info, file_name),
                    sheet_name,
                    self.main_window.options.get('sheet_name_rule', 'OriginalBoth'),
                    output_workbook.sheetnames,
                )
                output_sheet = output_workbook.create_sheet(title=new_sheet_name)
                self._copy_sheet_data(source_sheet, output_sheet, file_name=file_name)
                merged_count += 1
            except Exception as exc:
                self._log(f"시트 복사 오류 {item}: {exc}")
                sheet_errors.append(item)

            self._progress(int((i + 1) / total_sheets * 100))

        if sheet_errors:
            preview = ", ".join(sheet_errors[:3])
            if len(sheet_errors) > 3:
                preview += ", ..."
            raise RuntimeError(f"표준 엔진에서 {len(sheet_errors)}개 항목 병합 실패: {preview}")
        if merged_count == 0:
            raise RuntimeError("표준 엔진으로 병합할 수 있는 시트가 없습니다.")

        self._perform_sheet_trim(output_workbook)
        output_workbook.save(save_path)

    def merge_horizontally(self, sheets_to_merge, save_path):
        self._merge_by_axis(sheets_to_merge, save_path, 'horizontal')

    def merge_vertically(self, sheets_to_merge, save_path):
        self._merge_by_axis(sheets_to_merge, save_path, 'vertical')

    def _merge_by_axis(self, sheets_to_merge, save_path, axis):
        self._require_openpyxl()

        output_workbook = openpyxl.Workbook()
        output_sheet = output_workbook.active
        output_sheet.title = "Merged_Sheet"

        if self.main_window.options['only_value_copy']:
            self._log("수식을 값으로 변환하며 병합을 시작합니다...")

        total_sheets = len(sheets_to_merge)
        last_pos = 0
        sheet_errors = []
        merged_count = 0
        for i, item in enumerate(sheets_to_merge):
            file_name, sheet_name = item.split('/', 1)
            self._status(f'{item} 병합 중...')

            file_path = self.main_window.file_info.get(file_name, {}).get('processed_path')
            if not file_path:
                self._log(f"파일을 찾을 수 없습니다: {file_name}")
                sheet_errors.append(item)
                continue

            try:
                source_workbook = self.main_window.file_handler._open_workbook(
                    file_path, file_name, data_only=self.main_window.options['only_value_copy']
                )
                if not source_workbook:
                    continue

                if file_path.lower().endswith(('.xlsx', '.xlsm', '.xlsb', '.csv')):
                    source_sheet = source_workbook[sheet_name]
                else:
                    source_sheet = source_workbook.sheet_by_name(sheet_name)

                if axis == 'horizontal':
                    self._copy_sheet_data(source_sheet, output_sheet, start_col=last_pos + 1, file_name=file_name)
                    last_pos = output_sheet.max_column
                else:
                    self._copy_sheet_data(source_sheet, output_sheet, start_row=last_pos + 1, file_name=file_name)
                    last_pos = output_sheet.max_row
                merged_count += 1
            except Exception as exc:
                self._log(f"시트 병합 오류 {item}: {exc}")
                sheet_errors.append(item)

            self._progress(int((i + 1) / total_sheets * 100))

        if sheet_errors:
            preview = ", ".join(sheet_errors[:3])
            if len(sheet_errors) > 3:
                preview += ", ..."
            raise RuntimeError(f"표준 엔진에서 {len(sheet_errors)}개 항목 병합 실패: {preview}")
        if merged_count == 0:
            raise RuntimeError("표준 엔진으로 병합할 수 있는 시트가 없습니다.")

        self._perform_sheet_trim(output_workbook)
        output_workbook.save(save_path)

    def _copy_sheet_data(self, source_sheet, output_sheet, start_row=1, start_col=1, file_name=""):
        if openpyxl is not None and isinstance(source_sheet, openpyxl.worksheet.worksheet.Worksheet):
            for row in source_sheet.iter_rows():
                for cell in row:
                    new_cell = output_sheet.cell(row=cell.row + start_row - 1, column=cell.column + start_col - 1)
                    new_cell.value = cell.value
                    if cell.has_style:
                        new_cell.font = copy(cell.font)
                        new_cell.border = copy(cell.border)
                        new_cell.fill = copy(cell.fill)
                        new_cell.number_format = cell.number_format
                        new_cell.protection = copy(cell.protection)
                        new_cell.alignment = copy(cell.alignment)
                    if cell.hyperlink:
                        new_cell._hyperlink = copy(cell.hyperlink)
                    if cell.comment:
                        new_cell.comment = copy(cell.comment)

            for merged_range in source_sheet.merged_cells.ranges:
                output_sheet.merge_cells(
                    start_row=merged_range.min_row + start_row - 1,
                    start_column=merged_range.min_col + start_col - 1,
                    end_row=merged_range.max_row + start_row - 1,
                    end_column=merged_range.max_col + start_col - 1,
                )

            for col_letter, dimension in source_sheet.column_dimensions.items():
                target_dimension = output_sheet.column_dimensions[col_letter]
                target_dimension.width = dimension.width
                target_dimension.hidden = dimension.hidden

            for row_index, dimension in source_sheet.row_dimensions.items():
                target_dimension = output_sheet.row_dimensions[row_index + start_row - 1]
                target_dimension.height = dimension.height
                target_dimension.hidden = dimension.hidden

        elif xlrd is not None and isinstance(source_sheet, xlrd.sheet.Sheet):
            self._log(f".xls 파일({file_name}/{source_sheet.name})의 서식은 일부만 지원됩니다.")
            for row_idx in range(source_sheet.nrows):
                for col_idx in range(source_sheet.ncols):
                    cell_value = source_sheet.cell_value(row_idx, col_idx)
                    output_sheet.cell(row=row_idx + start_row, column=col_idx + start_col).value = cell_value

    def _perform_sheet_trim(self, workbook):
        sheet_trim_value = self.main_window.options.get('sheet_trim_value', 0)
        if sheet_trim_value <= 0:
            return

        trim_rows = self.main_window.options.get('sheet_trim_rows', False)
        trim_cols = self.main_window.options.get('sheet_trim_cols', False)
        if not trim_rows and not trim_cols:
            return

        self._log("시트 정리(SheetTrim) 기능 수행 중...")

        for worksheet in workbook.worksheets:
            if trim_rows:
                empty_row_indices = [
                    i for i in range(1, worksheet.max_row + 1)
                    if all(c.value is None or str(c.value).strip() == '' for c in worksheet[i])
                ]
                for start, count in reversed(self._build_trim_blocks(empty_row_indices, sheet_trim_value)):
                    worksheet.delete_rows(start, count)

            if trim_cols:
                empty_col_indices = [
                    i for i, col in enumerate(worksheet.iter_cols(), 1)
                    if all(cell.value is None or str(cell.value).strip() == "" for cell in col)
                ]
                for start, count in reversed(self._build_trim_blocks(empty_col_indices, sheet_trim_value)):
                    worksheet.delete_cols(start, count)

        self._log("시트 정리(SheetTrim) 완료.")

    def _build_trim_blocks(self, indexes, threshold):
        if not indexes:
            return []

        blocks = []
        start = indexes[0]
        prev = indexes[0]
        for current in indexes[1:]:
            if current != prev + 1:
                length = prev - start + 1
                if length >= threshold:
                    blocks.append((start, length))
                start = current
            prev = current

        length = prev - start + 1
        if length >= threshold:
            blocks.append((start, length))
        return blocks
