import sys
import os
import re
from PySide6.QtWidgets import QApplication, QMainWindow, QFileDialog, QListView, QAbstractItemView, QDialog
from PySide6.QtCore import (QCoreApplication, QStringListModel, Qt, QMimeData, QEvent, QPoint)
from PySide6.QtGui import QMouseEvent, QDrag, QKeySequence, QAction

from main_ui import Ui_MainWindow
from password_ui import Ui_Dialog as Ui_PasswordDialog
from globalpassword_ui import Ui_Dialog as Ui_GlobalPasswordDialog
from encryption_ui import Ui_Dialog as Ui_EncryptionDialog
from options_ui import Ui_Dialog as Ui_OptionsDialog

import openpyxl
import xlrd
import msoffcrypto
import configparser
from cryptography.fernet import Fernet
import io
import sys
import tempfile
import subprocess

if sys.platform == 'win32':
    try:
        import win32com.client as win32
    except ImportError:
        print("pywin32 is not installed. Please install it to use the .xls conversion feature on Windows.")
        win32 = None
else:
    win32 = None


class PasswordDialog(QDialog, Ui_PasswordDialog):
    def __init__(self, file_name, parent=None):
        super().__init__(parent)
        self.setupUi(self)
        self.textEditOpenFile.setText(file_name)
        self.lineEditKeepPassword.setFocus() # Set focus to the password input field
        self.btnStop.clicked.connect(self.on_stop_clicked)
        self.stopped = False

    def on_stop_clicked(self):
        self.stopped = True
        self.reject()

class GlobalPasswordDialog(QDialog, Ui_GlobalPasswordDialog):
    def __init__(self, parent=None):
        super().__init__(parent)
        self.setupUi(self)

class EncryptionDialog(QDialog, Ui_EncryptionDialog):
    def __init__(self, parent=None):
        super().__init__(parent)
        self.setupUi(self)

class OptionsDialog(QDialog, Ui_OptionsDialog):
    def __init__(self, parent=None, current_options=None):
        super().__init__(parent)
        self.setupUi(self)
        if current_options:
            self.set_options(current_options)

    def set_options(self, options):
        # Set merge type
        if options.get('merge_type') == 'Sheet':
            self.radioButtonSheet.setChecked(True)
        elif options.get('merge_type') == 'Horizontal':
            self.radioButtonHorizontal.setChecked(True)
        elif options.get('merge_type') == 'Vertical':
            self.radioButtonVertical.setChecked(True)

        # Set sheet name rule
        if options.get('sheet_name_rule') == 'OriginalBoth':
            self.radioButtonOriginalBoth.setChecked(True)
        elif options.get('sheet_name_rule') == 'OriginalSheet':
            self.radioButtonOriginalSheet.setChecked(True)

        # Set SheetTrim options
        self.spinBoxEmpty.setValue(options.get('sheet_trim_value', 0))
        self.checkBoxEmptyrow.setChecked(options.get('sheet_trim_rows', False))
        self.checkBoxEmptyColumn.setChecked(options.get('sheet_trim_cols', False))

    def get_options(self):
        options = {}
        if self.radioButtonSheet.isChecked():
            options['merge_type'] = 'Sheet'
        elif self.radioButtonHorizontal.isChecked():
            options['merge_type'] = 'Horizontal'
        elif self.radioButtonVertical.isChecked():
            options['merge_type'] = 'Vertical'

        if self.radioButtonOriginalBoth.isChecked():
            options['sheet_name_rule'] = 'OriginalBoth'
        elif self.radioButtonOriginalSheet.isChecked():
            options['sheet_name_rule'] = 'OriginalSheet'

        options['sheet_trim_value'] = self.spinBoxEmpty.value()
        options['sheet_trim_rows'] = self.checkBoxEmptyrow.isChecked()
        options['sheet_trim_cols'] = self.checkBoxEmptyColumn.isChecked()
        return options

class MainWindow(QMainWindow, Ui_MainWindow):
    def __init__(self):
        super().__init__()
        self.setupUi(self)
        self.setWindowTitle("Excel Merger")

        # Options state
        self.options = {
            'merge_type': 'Sheet',
            'sheet_name_rule': 'OriginalBoth',
            'sheet_trim_value': 0,
            'sheet_trim_rows': False,
            'sheet_trim_cols': False,
            'only_value_copy': False # New option for copying only values
        }

        # Settings and Encryption
        self.settings_file = 'config.ini'
        self.key = self.load_or_create_key()
        self.cipher = Fernet(self.key)
        self.global_password = ""
        self.output_encryption_password = ""
        self.use_global_password = False
        self.encrypt_output = False
        self.stop_asking_for_passwords = False
        self.debug_mode = False

        # File paths management
        self.file_info = {}
        self.file_passwords = {}
        self.temp_files = []

        # Initialize models for list views
        self.file_list_model = QStringListModel()
        self.listFileAdded.setModel(self.file_list_model)

        self.sheet_list_model = QStringListModel()
        self.listSheetInFile.setModel(self.sheet_list_model)

        self.merge_list_model = QStringListModel()
        self.listSheetToMerge.setModel(self.merge_list_model)

        # Store the currently selected file path
        self.current_selected_file_path = ""

        # Set properties for list views
        self.listFileAdded.setDragDropMode(QAbstractItemView.DragDropMode.InternalMove)
        self.listFileAdded.setSelectionMode(QAbstractItemView.SelectionMode.ExtendedSelection)
        self.listFileAdded.setDefaultDropAction(Qt.DropAction.MoveAction)
        self.listFileAdded.setAcceptDrops(True) # Allow dropping files

        self.listSheetInFile.setDragEnabled(True)
        self.listSheetInFile.setSelectionMode(QAbstractItemView.SelectionMode.ExtendedSelection)




        self.listFileAdded.setDragDropOverwriteMode(False)

        # Connect signals to slots
        self.actionAddExcelFile.triggered.connect(self.add_excel_file)
        self.listFileAdded.clicked.connect(self.on_file_selected)
        self.listFileAdded.doubleClicked.connect(self.remove_selected_files)

        self.btnSheetToMergeAdd.clicked.connect(self.add_sheet_to_merge)
        self.btnSheetToMergeRemove.clicked.connect(self.remove_sheet_from_merge)
        self.listSheetInFile.doubleClicked.connect(self.add_sheet_to_merge)
        self.listSheetToMerge.doubleClicked.connect(self.remove_sheet_from_merge)

        self.btnBrowsePath.clicked.connect(self.browse_save_path)
        self.btnOpenPath.clicked.connect(self.open_save_path_directory) # New connection
        self.actionSetSavePath.triggered.connect(self.browse_save_path)

        # Radio button connections for sheet selection
        self.radioButtonAll.toggled.connect(self.update_sheet_selection_mode)
        self.radioButtonSpecific.toggled.connect(self.update_sheet_selection_mode)
        self.radioButtonChoice.toggled.connect(self.update_sheet_selection_mode)
        self.lineEditSheetSpecific.textChanged.connect(self.populate_specific_sheets)


        # Password and Encryption actions
        self.menu_2.addAction(self.actionSetGlobalPassword)
        self.menu_2.addAction(self.actionSetOutputEncryption)
        self.menu_2.addAction(self.actionOptions)


        self.actionActivateDebugMode.setObjectName(u"actionActivateDebugMode")
        self.actionOptions.setShortcut(QCoreApplication.translate("MainWindow", u"Alt+O", None))
#endif // QT_CONFIG(shortcut)



        # Set initial state
        self.radioButtonChoice.setChecked(True)
        self.update_sheet_selection_mode()

        # Install event filter for key presses
        self.listFileAdded.installEventFilter(self)
        self.listFileAdded.viewport().installEventFilter(self)
        self.listSheetToMerge.installEventFilter(self)
        self.listSheetToMerge.viewport().installEventFilter(self)
        self.listSheetInFile.viewport().installEventFilter(self)
        self.listSheetInFile.installEventFilter(self)
        
        # Drag and Drop
        self.setAcceptDrops(False)
        self.drag_start_position = None


        self.actionActivateDebugMode.toggled.connect(self.on_debug_mode_toggled)

        # Connect start button
        self.btnStart.clicked.connect(self.start_merge)

        # Load settings and log initial state
        self.load_settings()

        # Connect checkBoxOnlyValue and set initial state
        self.checkBoxOnlyValue.setChecked(self.options['only_value_copy'])
        self.checkBoxOnlyValue.toggled.connect(self.on_only_value_copy_toggled)

    def closeEvent(self, event):
        # Clean up temporary files
        self.save_settings()
        for temp_file in self.temp_files:
            try:
                os.remove(temp_file)
                self.txtLogOutput.append(f"임시 파일 삭제: {temp_file}")
            except OSError as e:
                self.txtLogOutput.append(f"임시 파일 삭제 오류 {temp_file}: {e}")
        super().closeEvent(event)

    def perform_manual_move(self, list_view, model, event):
        # This function manually handles the reordering of items in a model.
        if self.debug_mode:
            log_msg = f"--- D&D Debug ---\n"
            log_msg += f"Event: {event.type()} on {list_view.objectName()}\n"
            log_msg += f"Pos: {event.pos()}\n"
            indicator_pos = list_view.dropIndicatorPosition()
            index_at_pos = list_view.indexAt(event.pos())
            log_msg += f"Indicator: {indicator_pos}\n"
            log_msg += f"Index @ Pos: {index_at_pos.row()}\n"
            selected_rows = [index.row() for index in list_view.selectedIndexes()]
            log_msg += f"Selected: {selected_rows}"
            self.txtLogOutput.append(log_msg)

        if event.type() == QEvent.Type.Drop:
            if self.debug_mode:
                self.txtLogOutput.append("-> Drop detected. Applying manual move.")
            
            dest_index = list_view.indexAt(event.pos())
            dest_row = dest_index.row()

            if list_view.dropIndicatorPosition() == QAbstractItemView.DropIndicatorPosition.BelowItem:
                dest_row += 1

            if dest_row == -1:
                dest_row = model.rowCount()

            source_indexes = list_view.selectedIndexes()
            source_rows = sorted([index.row() for index in source_indexes])
            source_data = [model.stringList()[row] for row in source_rows]

            if self.debug_mode:
                self.txtLogOutput.append(f"-> Source Rows: {source_rows} | Source Data: {source_data}")
                self.txtLogOutput.append(f"-> Initial Dest Row: {dest_row}")

            data_list = model.stringList()
            
            for row in reversed(source_rows):
                data_list.pop(row)

            offset = 0
            for row in source_rows:
                if row < dest_row:
                    offset += 1
            dest_row -= offset
            
            if self.debug_mode:
                self.txtLogOutput.append(f"-> Adjusted Dest Row: {dest_row}")

            for item in source_data:
                data_list.insert(dest_row, item)
                dest_row += 1
            
            model.setStringList(data_list)
            if self.debug_mode:
                self.txtLogOutput.append("-> Manual move complete.")
            
            event.accept()
            return True

        event.setDropAction(Qt.DropAction.MoveAction)
        return False

    def convert_xls_to_xlsx_win32(self, xls_path):
        if not win32:
            self.txtLogOutput.append("pywin32가 설치되지 않아 .xls 변환을 건너뜁니다.")
            return None
        excel = None # Initialize excel to None
        try:
            excel = win32.gencache.EnsureDispatch('Excel.Application')
            excel.Visible = False # Keep it hidden
            
            file_name = os.path.basename(xls_path)
            password = self.file_passwords.get(file_name)

            if password:
                wb = excel.Workbooks.Open(os.path.abspath(xls_path), UpdateLinks=0, Password=password)
            else:
                wb = excel.Workbooks.Open(os.path.abspath(xls_path), UpdateLinks=0)
            
            # Create a temporary file path for the xlsx file
            fd, xlsx_path = tempfile.mkstemp(suffix='.xlsx', prefix='excelmerger_')
            os.close(fd)

            excel.DisplayAlerts = False
            wb.SaveAs(xlsx_path, FileFormat=51) # 51 is for xlsx format
            excel.DisplayAlerts = True
            wb.Close()
            
            self.txtLogOutput.append(f".xls 파일을 .xlsx로 변환: {os.path.basename(xls_path)} -> {os.path.basename(xlsx_path)}")
            self.temp_files.append(xlsx_path)
            return xlsx_path
        except Exception as e:
            self.txtLogOutput.append(f".xls to .xlsx 변환 오류: {e}")
            return None
        finally:
            if excel:
                excel.DisplayAlerts = False
                excel.Application.Quit()

    def load_or_create_key(self):
        key_file = 'secret.key'
        if os.path.exists(key_file):
            with open(key_file, 'rb') as f:
                return f.read()
        else:
            key = Fernet.generate_key()
            with open(key_file, 'wb') as f:
                f.write(key)
            return key

    def load_settings(self):
        config = configparser.ConfigParser()
        if os.path.exists(self.settings_file):
            config.read(self.settings_file)
            if 'Passwords' in config:
                if 'global_password' in config['Passwords'] and config['Passwords']['global_password']:
                    try:
                        decrypted_pass = self.cipher.decrypt(config['Passwords']['global_password'].encode()).decode()
                        self.global_password = decrypted_pass
                    except:
                        self.global_password = ""
                self.use_global_password = config['Passwords'].getboolean('use_global_password', False)

                if 'output_encryption_password' in config['Passwords'] and config['Passwords']['output_encryption_password']:
                    try:
                        decrypted_pass = self.cipher.decrypt(config['Passwords']['output_encryption_password'].encode()).decode()
                        self.output_encryption_password = decrypted_pass
                    except:
                        self.output_encryption_password = ""
                self.encrypt_output = config['Passwords'].getboolean('encrypt_output', False)

        if 'Options' in config:
            self.options['merge_type'] = config['Options'].get('merge_type', 'Sheet')
            self.options['sheet_name_rule'] = config['Options'].get('sheet_name_rule', 'OriginalBoth')
            self.options['sheet_trim_value'] = config['Options'].getint('sheet_trim_value', 0)
            self.options['sheet_trim_rows'] = config['Options'].getboolean('sheet_trim_rows', False)
            self.options['sheet_trim_cols'] = config['Options'].getboolean('sheet_trim_cols', False)
            self.options['only_value_copy'] = config['Options'].getboolean('only_value_copy', False) # Load new option
            self.debug_mode = config['Options'].getboolean('debug_mode', False)
            self.actionActivateDebugMode.setChecked(self.debug_mode)

        if 'Paths' in config and 'last_save_path' in config['Paths']:
            self.lineEditSavePath.setText(config['Paths']['last_save_path'])

        self.txtLogOutput.append(f"출력파일 암호화: {'활성' if self.encrypt_output else '비활성'}")
        self.txtLogOutput.append(f"전역 비밀번호: {'설정됨' if self.use_global_password and self.global_password else '설정 안됨'}")

    def save_settings(self):
        config = configparser.ConfigParser()
        config.read(self.settings_file)

        if not config.has_section('Passwords'):
            config.add_section('Passwords')
        encrypted_global_pass = self.cipher.encrypt(self.global_password.encode()).decode() if self.global_password else ""
        config.set('Passwords', 'global_password', encrypted_global_pass)
        config.set('Passwords', 'use_global_password', str(self.use_global_password))
        encrypted_output_pass = self.cipher.encrypt(self.output_encryption_password.encode()).decode() if self.output_encryption_password else ""
        config.set('Passwords', 'output_encryption_password', encrypted_output_pass)
        config.set('Passwords', 'encrypt_output', str(self.encrypt_output))

        if not config.has_section('Options'):
            config.add_section('Options')
        for key, value in self.options.items():
            config.set('Options', key, str(value))
        config.set('Options', 'debug_mode', str(self.debug_mode))

        if not config.has_section('Paths'):
            config.add_section('Paths')
        config.set('Paths', 'last_save_path', self.lineEditSavePath.text())

        with open(self.settings_file, 'w') as configfile:
            config.write(configfile)

    def open_global_password_dialog(self):
        dialog = GlobalPasswordDialog(self)
        dialog.chkGlobalPassword.setChecked(self.use_global_password)
        dialog.lineEditGlobalPassword.setText(self.global_password)
        dialog.lineEditGlobalPassword.setEnabled(self.use_global_password)

        dialog.chkGlobalPassword.toggled.connect(dialog.lineEditGlobalPassword.setEnabled)

        if dialog.exec():
            self.use_global_password = dialog.chkGlobalPassword.isChecked()
            self.global_password = dialog.lineEditGlobalPassword.text() if self.use_global_password else ""
            self.save_settings()
            self.txtLogOutput.append("전역 비밀번호 설정이 업데이트되었습니다.")

    def open_encryption_dialog(self):
        dialog = EncryptionDialog(self)
        dialog.chkEnablePassword.setChecked(self.encrypt_output)
        dialog.lineEditPassword.setText(self.output_encryption_password)
        dialog.lineEditPassword.setEnabled(self.encrypt_output)

        dialog.chkEnablePassword.toggled.connect(dialog.lineEditPassword.setEnabled)

        if dialog.exec():
            self.encrypt_output = dialog.chkEnablePassword.isChecked()
            self.output_encryption_password = dialog.lineEditPassword.text() if self.encrypt_output else ""
            self.save_settings()
            self.txtLogOutput.append("출력 파일 암호화 설정이 업데이트되었습니다.")

    def open_options_dialog(self):
        # Create a temporary dictionary with only the options managed by OptionsDialog
        options_for_dialog = {
            'merge_type': self.options.get('merge_type', 'Sheet'),
            'sheet_name_rule': self.options.get('sheet_name_rule', 'OriginalBoth'),
            'sheet_trim_value': self.options.get('sheet_trim_value', 0),
            'sheet_trim_rows': self.options.get('sheet_trim_rows', False),
            'sheet_trim_cols': self.options.get('sheet_trim_cols', False)
        }
        dialog = OptionsDialog(self, current_options=options_for_dialog)
        if dialog.exec():
            # Update only the options managed by OptionsDialog
            updated_options = dialog.get_options()
            self.options['merge_type'] = updated_options.get('merge_type', 'Sheet')
            self.options['sheet_name_rule'] = updated_options.get('sheet_name_rule', 'OriginalBoth')
            self.options['sheet_trim_value'] = updated_options.get('sheet_trim_value', 0)
            self.options['sheet_trim_rows'] = updated_options.get('sheet_trim_rows', False)
            self.options['sheet_trim_cols'] = updated_options.get('sheet_trim_cols', False)
            self.txtLogOutput.append("옵션이 업데이트되었습니다.")
            self.save_settings()

    def on_only_value_copy_toggled(self, checked):
        self.options['only_value_copy'] = checked
        self.save_settings()

    def on_debug_mode_toggled(self, checked):
        self.debug_mode = checked
        self.txtLogOutput.append(f"디버그 모드: {'활성' if checked else '비활성'}")
        self.save_settings()

    def _open_workbook(self, file_path, file_name, data_only=False): # Add data_only flag
        try:
            if file_path.endswith('.xlsx'):
                return openpyxl.load_workbook(file_path, read_only=False, data_only=data_only) # Use data_only flag
            elif file_path.endswith('.xls'):
                # xlrd does not have a direct equivalent of data_only=True when opening.
                # It always reads the calculated values.
                return xlrd.open_workbook(file_path, formatting_info=True)
        except Exception as e:
            self.txtLogOutput.append(f"파일 열기 오류 {file_name}: {e}")
            return None
        return None

    def get_sheet_names(self, file_path):
        file_name = os.path.basename(file_path)
        workbook = self._open_workbook(file_path, file_name)
        
        original_file_path = file_path # Keep track of the original path
        processed_file_path = original_file_path # Initialize processed_file_path

        if workbook is None:
            # If _open_workbook failed, try to handle it as an encrypted file
            decrypted_temp_path = self.handle_encrypted_file(original_file_path)
            if decrypted_temp_path:
                # If successfully decrypted, try to open the workbook again with the temp file
                workbook = self._open_workbook(decrypted_temp_path, file_name)
                processed_file_path = decrypted_temp_path # Update processed_file_path
                if workbook is None:
                    self.txtLogOutput.append(f"암호 해독된 파일 열기 실패: {file_name}")
                    return None, None # Return None for both
            else:
                # If handle_encrypted_file also failed or user cancelled
                self.txtLogOutput.append(f"파일을 열 수 없습니다 (암호화 문제 또는 사용자 취소): {file_name}")
                return None, None # Return None for both

        # If we have a workbook (either original or decrypted temp), get sheet names
        if workbook:
            try:
                if processed_file_path.endswith('.xlsx'): # Use processed_file_path here
                    return workbook.sheetnames, processed_file_path
                elif processed_file_path.endswith('.xls'): # Use processed_file_path here
                    return workbook.sheet_names(), processed_file_path
            except Exception as e:
                self.txtLogOutput.append(f"시트 이름 가져오기 오류 {file_name}: {e}")
                return None, None # Return None for both
        return None, None # Should not be reached if workbook is valid

    def eventFilter(self, source, event):
        # File drop on listFileAdded
        if source == self.listFileAdded and event.type() in (QEvent.Type.DragEnter, QEvent.Type.Drop):
            if event.mimeData().hasUrls():
                if event.type() == QEvent.Type.DragEnter:
                    event.acceptProposedAction()
                else:
                    files = [url.toLocalFile() for url in event.mimeData().urls()]
                    self.add_files(files)
                    event.acceptProposedAction()
                return True

        # Internal move for listFileAdded
        elif source == self.listFileAdded.viewport() and event.type() in (QEvent.Type.DragEnter, QEvent.Type.DragMove, QEvent.Type.Drop):
            # Internal move logic is handled by the helper function
            if self.perform_manual_move(self.listFileAdded, self.file_list_model, event):
                return True
            else:
                return False

        # Drag start from listSheetInFile
        if source == self.listSheetInFile.viewport():
            if event.type() == QEvent.Type.MouseButtonPress:
                self.drag_start_position = event.pos()
            elif event.type() == QEvent.Type.MouseMove and self.drag_start_position:
                if (event.pos() - self.drag_start_position).manhattanLength() > QApplication.startDragDistance():
                    self.perform_drag_sheet_in_file()
        
        # Drag/Drop on listSheetToMerge's viewport
        elif source == self.listSheetToMerge.viewport() and event.type() in (QEvent.Type.DragEnter, QEvent.Type.DragMove, QEvent.Type.Drop):
            if event.mimeData().hasFormat("application/x-sheet-data"):
                # External drop from listSheetInFile
                if event.type() != QEvent.Type.Drop:
                    event.acceptProposedAction()
                else:
                    self.handle_sheet_drop(event)
                return True # We handled it
            else:
                # Internal move for listSheetToMerge
                if self.perform_manual_move(self.listSheetToMerge, self.merge_list_model, event):
                    return True
                else:
                    return False

        # Key presses
        elif event.type() == QEvent.Type.KeyPress:
            if event.key() == Qt.Key.Key_Delete:
                if source == self.listFileAdded:
                    self.remove_selected_files()
                    return True
                elif source == self.listSheetToMerge:
                    self.remove_sheet_from_merge()
                    return True
            elif event.key() in (Qt.Key.Key_Return, Qt.Key.Key_Enter):
                if source == self.listSheetInFile:
                    self.add_sheet_to_merge()
                    return True

        return super().eventFilter(source, event)



    def dragMoveEvent(self, event):
        if event.mimeData().hasFormat("application/x-sheet-data"):
            event.acceptProposedAction()



    def perform_drag_sheet_in_file(self):
        indexes = self.listSheetInFile.selectedIndexes()
        if not indexes:
            return

        mime_data = QMimeData()
        sheet_data = []

        file_path = self.current_selected_file_path
        file_name = os.path.basename(file_path)

        for index in indexes:
            sheet_name = self.sheet_list_model.data(index, Qt.ItemDataRole.DisplayRole)
            sheet_data.append(f"{file_name}|{sheet_name}")
        
        mime_data.setText("\n".join(sheet_data))
        mime_data.setData("application/x-sheet-data", mime_data.text().encode())

        drag = QDrag(self.listSheetInFile)
        drag.setMimeData(mime_data)
        drag.exec(Qt.DropAction.CopyAction)

        self.drag_start_position = None

    def handle_sheet_drop(self, event):
        sheet_data_raw = event.mimeData().data("application/x-sheet-data").data().decode()
        sheet_items = sheet_data_raw.split('\n')
        
        current_merge_list = self.merge_list_model.stringList()
        for item_str in sheet_items:
            file_name, sheet_name = item_str.split('|')
            formatted_item = f"{file_name}/{sheet_name}"
            if formatted_item not in current_merge_list:
                current_merge_list.append(formatted_item)
        self.merge_list_model.setStringList(current_merge_list)
        event.acceptProposedAction()

    def add_files(self, files):
        for file in files:
            basename = os.path.basename(file)
            if basename in self.file_info: # Skip duplicates
                continue

            if file.endswith('.xlsx') or file.endswith('.xls'):
                sheet_names, processed_file_path = self.get_sheet_names(file)

                if sheet_names is not None and processed_file_path is not None:
                    self.file_info[basename] = {
                        'original_path': file,
                        'processed_path': processed_file_path, # Use the processed path
                        'sheets': sheet_names
                    }
                else:
                    self.txtLogOutput.append(f"파일을 열 수 없어 목록에서 제외합니다: {basename}")

        self.file_list_model.setStringList(list(self.file_info.keys()))


    def add_excel_file(self):
        self.txtLogOutput.append("add_excel_file triggered")
        files, _ = QFileDialog.getOpenFileNames(self, "ADD EXCEL FILES DIALOG", "", "Excel Files (*.xlsx *.xls)")
        if files:
            self.add_files(files)

    def remove_selected_files(self):
        indexes = self.listFileAdded.selectedIndexes()
        if not indexes:
            return
        
        basenames_to_remove = [self.file_list_model.data(index, 0) for index in indexes]

        for basename in basenames_to_remove:
            if basename in self.file_info:
                del self.file_info[basename]
            if basename in self.file_passwords:
                del self.file_passwords[basename]
        
        self.file_list_model.setStringList(list(self.file_info.keys()))
        self.sheet_list_model.setStringList([])
        self.merge_list_model.setStringList([])


    def on_file_selected(self, index):
        basename = self.file_list_model.data(index, 0)
        info = self.file_info.get(basename)
        if info:
            self.current_selected_file_path = info['processed_path']
            self.load_sheets(basename)

    def load_sheets(self, basename):
        info = self.file_info.get(basename)
        sheet_names = info.get('sheets') if info else None
        self.sheet_list_model.setStringList(sheet_names if sheet_names else [])


    def add_sheet_to_merge(self):
        if not self.radioButtonChoice.isChecked():
            return
        indexes = self.listSheetInFile.selectedIndexes()
        if not indexes:
            return

        current_merge_list = self.merge_list_model.stringList()
        file_name = os.path.basename(self.current_selected_file_path)

        for index in indexes:
            sheet_name = self.sheet_list_model.data(index, Qt.ItemDataRole.DisplayRole)
            formatted_item = f"{file_name}/{sheet_name}"
            if formatted_item not in current_merge_list:
                current_merge_list.append(formatted_item)
        
        self.merge_list_model.setStringList(current_merge_list)

    def remove_sheet_from_merge(self):
        if not self.radioButtonChoice.isChecked():
            return
        indexes = self.listSheetToMerge.selectedIndexes()
        if not indexes:
            return

        rows_to_remove = sorted([index.row() for index in indexes], reverse=True)
        current_list = self.merge_list_model.stringList()

        for row in rows_to_remove:
            del current_list[row]
        
        self.merge_list_model.setStringList(current_list)

    def browse_save_path(self):
        save_path, _ = QFileDialog.getSaveFileName(self, "Save Merged File", "", "Excel Files (*.xlsx)")
        if save_path:
            self.lineEditSavePath.setText(save_path)
            self.save_settings()

    def open_save_path_directory(self):
        path = self.lineEditSavePath.text()
        if not path:
            self.txtLogOutput.append("저장 경로가 지정되지 않았습니다.")
            return

        directory = os.path.dirname(path)
        if not os.path.isdir(directory):
            self.txtLogOutput.append(f"디렉토리를 찾을 수 없습니다: {directory}")
            return

        try:
            if sys.platform == 'win32':
                os.startfile(directory)
            elif sys.platform == 'darwin': # macOS
                subprocess.run(['open', directory])
            else: # Linux
                subprocess.run(['xdg-open', directory])
        except Exception as e:
            self.txtLogOutput.append(f"디렉토리를 열 수 없습니다: {e}")

    def update_sheet_selection_mode(self):
        is_choice_mode = self.radioButtonChoice.isChecked()

        # Keep lists enabled for scrolling, but control interaction
        self.listSheetInFile.setEnabled(True)
        self.listSheetToMerge.setEnabled(True)

        self.btnSheetToMergeAdd.setEnabled(is_choice_mode)
        self.btnSheetToMergeRemove.setEnabled(is_choice_mode)

        if is_choice_mode:
            self.listSheetInFile.setSelectionMode(QAbstractItemView.SelectionMode.ExtendedSelection)
            self.listSheetInFile.setDragEnabled(True)
            self.listSheetToMerge.setDragDropMode(QAbstractItemView.DragDropMode.InternalMove)
            self.listSheetToMerge.setSelectionMode(QAbstractItemView.SelectionMode.ExtendedSelection)
        else:
            self.listSheetInFile.clearSelection()
            self.listSheetInFile.setSelectionMode(QAbstractItemView.SelectionMode.NoSelection)
            self.listSheetInFile.setDragEnabled(False)
            
            self.listSheetToMerge.clearSelection()
            self.listSheetToMerge.setDragDropMode(QAbstractItemView.DragDropMode.NoDragDrop)
            self.listSheetToMerge.setSelectionMode(QAbstractItemView.SelectionMode.NoSelection)

        if self.radioButtonAll.isChecked():
            self.lineEditSheetSpecific.setEnabled(False)
            self.populate_all_sheets()
        elif self.radioButtonSpecific.isChecked():
            self.lineEditSheetSpecific.setEnabled(True)
            self.populate_specific_sheets()
        elif self.radioButtonChoice.isChecked():
            self.lineEditSheetSpecific.setEnabled(False)
            # In choice mode, the user manages the list, so we don't clear it automatically
            # unless it's the initial switch to this mode.
            # The logic in __init__ handles the initial state.
            pass

    def populate_all_sheets(self):
        self.stop_asking_for_passwords = False
        all_sheets_to_merge = []
        for file_name, info in self.file_info.items():
            sheet_names = info.get('sheets')
            if sheet_names:
                for sheet_name in sheet_names:
                    all_sheets_to_merge.append(f"{file_name}/{sheet_name}")

        self.merge_list_model.setStringList(all_sheets_to_merge)

    def populate_specific_sheets(self):
        self.stop_asking_for_passwords = False
        sheet_indices_str = self.lineEditSheetSpecific.text()
        if not sheet_indices_str:
            self.merge_list_model.setStringList([])
            return
            
        try:
            sheet_indices = [int(i.strip()) - 1 for i in sheet_indices_str.split(',') if i.strip().isdigit()]
        except:
            self.merge_list_model.setStringList([])
            return

        specific_sheets_to_merge = []
        for file_name, info in self.file_info.items():
            sheet_names = info.get('sheets')
            if sheet_names:
                for index in sheet_indices:
                    if 0 <= index < len(sheet_names):
                        specific_sheets_to_merge.append(f"{file_name}/{sheet_names[index]}")
        
        self.merge_list_model.setStringList(specific_sheets_to_merge)

    def handle_encrypted_file(self, file_path):
        if self.stop_asking_for_passwords:
            self.txtLogOutput.append(f"비밀번호 입력을 중단하여 파일을 건너뜁니다: {os.path.basename(file_path)}")
            if self.debug_mode:
                self.txtLogOutput.append(f"DEBUG: handle_encrypted_file returning None (stopped asking).") # ADD LOG
            return None

        self.txtLogOutput.append(f"암호화된 파일 감지: {os.path.basename(file_path)}")
        password = None
        decrypted_file_buffer = None
        basename = os.path.basename(file_path)
        temp_decrypted_path = None

        # Password check order: 1. File-specific, 2. Global, 3. User prompt
        if basename in self.file_passwords:
            try:
                self.txtLogOutput.append(f'{basename}에 대해 기억된 비밀번호로 열기 시도...')
                decrypted_file_buffer = io.BytesIO()
                with open(file_path, 'rb') as f:
                    office_file = msoffcrypto.OfficeFile(f)
                    office_file.load_key(password=self.file_passwords[basename])
                    office_file.decrypt(decrypted_file_buffer)
                password = self.file_passwords[basename]
                self.txtLogOutput.append("기억된 비밀번호로 열기 성공.")
            except Exception:
                self.txtLogOutput.append("기억된 비밀번호 실패.")
                decrypted_file_buffer = None

        if not password and self.use_global_password and self.global_password:
            try:
                self.txtLogOutput.append("전역 비밀번호로 열기 시도...")
                decrypted_file_buffer = io.BytesIO()
                with open(file_path, 'rb') as f:
                    office_file = msoffcrypto.OfficeFile(f)
                    office_file.load_key(password=self.global_password)
                    office_file.decrypt(decrypted_file_buffer)
                password = self.global_password
                self.file_passwords[basename] = self.global_password
                self.txtLogOutput.append("전역 비밀번호로 열기 성공.")
            except Exception as e:
                self.txtLogOutput.append(f"전역 비밀번호 실패: {e}")
                decrypted_file_buffer = None

        if not password:
            if self.debug_mode:
                self.txtLogOutput.append(f"DEBUG: Showing PasswordDialog for {basename}.")
            dialog = PasswordDialog(basename, self)
            result = dialog.exec()

            if dialog.stopped:
                self.stop_asking_for_passwords = True
                self.txtLogOutput.append("사용자가 중단하여 이후 암호 입력을 건너뜁니다.")
                if self.debug_mode:
                    self.txtLogOutput.append(f"DEBUG: handle_encrypted_file returning None (dialog stopped).")
                return None

            if result:
                user_password = dialog.lineEditKeepPassword.text()
                if user_password:
                    try:
                        self.txtLogOutput.append("사용자 입력 비밀번호로 열기 시도...")
                        decrypted_file_buffer = io.BytesIO()
                        with open(file_path, 'rb') as f:
                            office_file = msoffcrypto.OfficeFile(f)
                            office_file.load_key(password=user_password)
                            office_file.decrypt(decrypted_file_buffer)
                        password = user_password
                        self.txtLogOutput.append("사용자 입력 비밀번호로 열기 성공.")
                        self.file_passwords[basename] = user_password
                        if dialog.chkKeepPassword.isChecked():
                            self.global_password = user_password
                            self.use_global_password = True
                    except Exception as e:
                        self.txtLogOutput.append(f"사용자 입력 비밀번호 실패: {e}")
                        if self.debug_mode:
                            self.txtLogOutput.append(f"DEBUG: handle_encrypted_file returning None (user password failed).")
                        return None
                else:
                    self.txtLogOutput.append("비밀번호가 입력되지 않아 파일을 건너뜁니다.")
                    if self.debug_mode:
                        self.txtLogOutput.append(f"DEBUG: handle_encrypted_file returning None (empty user password).")
                    return None
            else:
                self.txtLogOutput.append("사용자가 취소하여 파일을 건너뜁니다.")
                if self.debug_mode:
                    self.txtLogOutput.append(f"DEBUG: handle_encrypted_file returning None (dialog cancelled).")
                return None

        if password and decrypted_file_buffer:
            fd, temp_decrypted_path = tempfile.mkstemp(suffix=os.path.splitext(file_path)[1], prefix='decrypted_')
            os.close(fd)
            with open(temp_decrypted_path, 'wb') as tmp_f:
                tmp_f.write(decrypted_file_buffer.getbuffer())
            self.temp_files.append(temp_decrypted_path)
            if self.debug_mode:
                self.txtLogOutput.append(f"DEBUG: handle_encrypted_file returning decrypted path: {temp_decrypted_path}")
            return temp_decrypted_path
        
        if self.debug_mode:
            self.txtLogOutput.append(f"DEBUG: handle_encrypted_file returning None (no password worked).")
        return None

    def start_merge(self):
        sheets_to_merge = self.merge_list_model.stringList()
        if not sheets_to_merge:
            self.txtLogOutput.append("병합할 시트가 없습니다.")
            return

        save_path = self.lineEditSavePath.text()
        if not save_path:
            self.txtLogOutput.append("저장 경로를 지정하세요.")
            return

        # Validate save path
        save_dir = os.path.dirname(save_path)
        if not os.path.isdir(save_dir):
            self.txtLogOutput.append(f"경고: 저장 경로의 디렉토리가 존재하지 않습니다: {save_dir}")
            return
        if not os.access(save_dir, os.W_OK):
            self.txtLogOutput.append(f"경고: 저장 경로에 쓸 수 있는 권한이 없습니다: {save_dir}")
            return

        self.progressBar.setValue(0)
        self.txtLogOutput.clear()
        if self.debug_mode:
            self.txtLogOutput.append(f"DEBUG: file_passwords at start of merge: {self.file_passwords}")

        try:
            merge_type = self.options.get('merge_type', 'Sheet')

            # Check if win32com is available and Excel is installed for high-quality merge
            use_win32_merge = False
            if sys.platform == 'win32' and win32:
                try:
                    # Attempt to dispatch Excel application to check if it's installed
                    excel_app_check = win32.gencache.EnsureDispatch('Excel.Application')
                    excel_app_check.Quit() # Quit immediately after checking
                    use_win32_merge = True
                except Exception as e:
                    self.txtLogOutput.append(f"Excel 애플리케이션을 찾을 수 없습니다. 고품질 병합을 건너뜁니다: {e}")
                    use_win32_merge = False

            if merge_type != 'Sheet' and use_win32_merge:
                self.txtLogOutput.append("고품질 병합을 위해 .xls 파일을 .xlsx로 변환합니다...")
                for basename, info in self.file_info.items():
                    if info['original_path'].endswith('.xls'):
                        new_path = self.convert_xls_to_xlsx_win32(info['original_path'])
                        if new_path:
                            self.file_info[basename]['processed_path'] = new_path

            if merge_type == 'Sheet':
                if use_win32_merge: # Use win32 merge if available, regardless of only_value_copy
                    self.merge_as_sheets_win32(sheets_to_merge, save_path)
                else:
                    self.merge_as_sheets(sheets_to_merge, save_path)
            elif merge_type == 'Horizontal':
                if use_win32_merge:
                    self.merge_horizontally_win32(sheets_to_merge, save_path)
                else:
                    self.merge_horizontally(sheets_to_merge, save_path)
            elif merge_type == 'Vertical':
                if use_win32_merge:
                    self.merge_vertically_win32(sheets_to_merge, save_path)
                else:
                    self.merge_vertically(sheets_to_merge, save_path)
            
            # Encrypt output file if needed
            if self.encrypt_output and self.output_encryption_password:
                self.txtLogOutput.append("출력 파일 암호화 중...")
                self.txtLogOutput.append(f"암호화 설정: {self.encrypt_output}, 암호 길이: {len(self.output_encryption_password)}")
                encrypted_file = io.BytesIO()
                with open(save_path, 'rb') as f:
                    office_file = msoffcrypto.OfficeFile(f)
                    office_file.encrypt(self.output_encryption_password, encrypted_file)
                
                with open(save_path, 'wb') as f:
                    f.write(encrypted_file.getbuffer())
                self.txtLogOutput.append("출력 파일 암호화 완료.")

            self.txtLogOutput.append(f"병합 완료: {save_path}")

        except Exception as e:
            self.txtLogOutput.append(f"병합 오류: {e}")
        finally:
            self.progressBar.setValue(100)
            # Open the save path directory and focus on the file
            try:
                if sys.platform == 'win32':
                    subprocess.run(['explorer', '/select,', os.path.abspath(save_path)])
                elif sys.platform == 'darwin': # macOS
                    subprocess.run(['open', '-R', os.path.abspath(save_path)])
                else: # Linux
                    subprocess.run(['xdg-open', os.path.dirname(os.path.abspath(save_path))])
            except Exception as e:
                self.txtLogOutput.append(f"저장 경로를 열 수 없습니다: {e}")

    def merge_as_sheets_win32(self, sheets_to_merge, save_path):
        excel = None
        try:
            excel = win32.gencache.EnsureDispatch('Excel.Application')
            excel.Visible = False
            merged_workbook = excel.Workbooks.Add()

            # The default workbook has one sheet, we need to know its name to delete it later
            default_sheet_name = merged_workbook.Worksheets(1).Name

            total_sheets = len(sheets_to_merge)
            for i, item in enumerate(sheets_to_merge):
                file_name, sheet_name = item.split('/', 1)
                self.lblCurrentFile.setText(f'{item} 병합 중 (고품질 모드)...')
                QApplication.processEvents()

                info = self.file_info.get(file_name)
                if not info:
                    self.txtLogOutput.append(f"파일 정보를 찾을 수 없습니다: {file_name}")
                    continue
                
                processed_path = os.path.abspath(info['processed_path'])
                
                try:
                    if password:
                        if self.debug_mode:
                            self.txtLogOutput.append(f"DEBUG: {file_name}에 기억된 비밀번호로 열기 시도 (Win32)...")
                        source_workbook = excel.Workbooks.Open(processed_path, UpdateLinks=0, Password=password)
                    else:
                        source_workbook = excel.Workbooks.Open(processed_path, UpdateLinks=0)
                    
                    source_sheet = source_workbook.Worksheets(sheet_name)
                    
                    # Copy sheet to the end of the merged workbook
                    source_sheet.Copy(After=merged_workbook.Worksheets(merged_workbook.Worksheets.Count))
                    
                    # Get the newly copied sheet
                    newly_copied_sheet = merged_workbook.Worksheets(merged_workbook.Worksheets.Count)
                    
                    source_workbook.Close(SaveChanges=False)
                except Exception as e:
                    self.txtLogOutput.append(f"시트 복사 오류 (win32) {item}: {e}")
                
                self.progressBar.setValue(int((i + 1) / total_sheets * 100))
                QApplication.processEvents()

            # Delete the default sheet that was created with the new workbook
            if merged_workbook.Worksheets.Count > 1:
                try:
                    merged_workbook.Worksheets(default_sheet_name).Delete()
                except:
                    # Ignore error if sheet is already gone or name is different
                    pass

            # If only_value_copy is enabled, convert all formulas to values in the merged workbook
            if self.options['only_value_copy']:
                if self.debug_mode:
                    self.txtLogOutput.append("DEBUG: 병합된 시트의 수식을 값으로 변환 중...")
                for ws in merged_workbook.Worksheets:
                    # Get the used range of the worksheet
                    used_range = ws.UsedRange
                    # Copy the used range
                    used_range.Copy()
                    # Paste special values to the same range
                    used_range.PasteSpecial(Paste=win32.constants.xlPasteValues)
                    # Clear the clipboard
                    excel.CutCopyMode = False

            # Suppress alerts to automatically overwrite existing files
            excel.DisplayAlerts = False # <--- Add this line
            merged_workbook.SaveAs(os.path.abspath(save_path))
            excel.DisplayAlerts = True # <--- Restore alerts after saving
            merged_workbook.Close(SaveChanges=False)

        except Exception as e:
            self.txtLogOutput.append(f"win32 병합 오류: {e}")
        finally:
            if excel:
                excel.DisplayAlerts = False
                excel.Application.Quit()

    def merge_horizontally_win32(self, sheets_to_merge, save_path):
        excel = None
        try:
            excel = win32.gencache.EnsureDispatch('Excel.Application')
            excel.Visible = False
            output_workbook = excel.Workbooks.Add()
            output_sheet = output_workbook.Worksheets(1)
            output_sheet.Name = "Merged_Sheet"

            total_sheets = len(sheets_to_merge)
            last_col = 0
            for i, item in enumerate(sheets_to_merge):
                file_name, sheet_name = item.split('/', 1)
                self.lblCurrentFile.setText(f'{item} 병합 중 (고품질 모드)...')
                QApplication.processEvents()

                info = self.file_info.get(file_name)
                if not info:
                    self.txtLogOutput.append(f"파일 정보를 찾을 수 없습니다: {file_name}")
                    continue
                
                processed_path = os.path.abspath(info['processed_path'])
                
                password = self.file_passwords.get(file_name)

                try:
                    if password:
                        source_workbook = excel.Workbooks.Open(processed_path, UpdateLinks=0, Password=password)
                    else:
                        source_workbook = excel.Workbooks.Open(processed_path, UpdateLinks=0)
                    
                    source_sheet = source_workbook.Worksheets(sheet_name)
                    
                    source_range = source_sheet.UsedRange
                    if source_range.Columns.Count > 0:
                        source_range.Copy()
                        
                        destination_range = output_sheet.Cells(1, last_col + 1)
                        output_sheet.Paste(Destination=destination_range)
                        
                        last_col += source_range.Columns.Count
                    
                    source_workbook.Close(SaveChanges=False)
                except Exception as e:
                    self.txtLogOutput.append(f"시트 병합 오류 (win32) {item}: {e}")
                
                self.progressBar.setValue(int((i + 1) / total_sheets * 100))
                QApplication.processEvents()

            if self.options['only_value_copy']:
                self.txtLogOutput.append("병합된 시트의 수식을 값으로 변환 중 (고품질 모드)...")
                used_range = output_sheet.UsedRange
                used_range.Copy()
                used_range.PasteSpecial(Paste=win32.constants.xlPasteValues)
                excel.CutCopyMode = False

            excel.DisplayAlerts = False
            output_workbook.SaveAs(os.path.abspath(save_path))
            excel.DisplayAlerts = True
            output_workbook.Close(SaveChanges=False)

        except Exception as e:
            self.txtLogOutput.append(f"win32 병합 오류: {e}")
        finally:
            if excel:
                excel.DisplayAlerts = False
                excel.Application.Quit()

    def merge_vertically_win32(self, sheets_to_merge, save_path):
        excel = None
        try:
            excel = win32.gencache.EnsureDispatch('Excel.Application')
            excel.Visible = False
            output_workbook = excel.Workbooks.Add()
            output_sheet = output_workbook.Worksheets(1)
            output_sheet.Name = "Merged_Sheet"

            total_sheets = len(sheets_to_merge)
            last_row = 0
            for i, item in enumerate(sheets_to_merge):
                file_name, sheet_name = item.split('/', 1)
                self.lblCurrentFile.setText(f'{item} 병합 중 (고품질 모드)...')
                QApplication.processEvents()

                info = self.file_info.get(file_name)
                if not info:
                    self.txtLogOutput.append(f"파일 정보를 찾을 수 없습니다: {file_name}")
                    continue
                
                processed_path = os.path.abspath(info['processed_path'])
                
                password = self.file_passwords.get(file_name)

                try:
                    if password:
                        if self.debug_mode:
                            self.txtLogOutput.append(f"DEBUG: {file_name}에 기억된 비밀번호로 열기 시도 (Win32)...")
                        source_workbook = excel.Workbooks.Open(processed_path, UpdateLinks=0, Password=password)
                    else:
                        source_workbook = excel.Workbooks.Open(processed_path, UpdateLinks=0)
                    
                    source_sheet = source_workbook.Worksheets(sheet_name)
                    
                    source_range = source_sheet.UsedRange
                    if source_range.Rows.Count > 0:
                        source_range.Copy()
                        
                        destination_range = output_sheet.Cells(last_row + 1, 1)
                        if self.options['only_value_copy']:
                            destination_range.PasteSpecial(Paste=win32.constants.xlPasteValues)
                            excel.CutCopyMode = False
                        else:
                            output_sheet.Paste(Destination=destination_range)
                        
                        last_row += source_range.Rows.Count
                    
                    source_workbook.Close(SaveChanges=False)
                except Exception as e:
                    self.txtLogOutput.append(f"시트 병합 오류 (win32) {item}: {e}")
                
                self.progressBar.setValue(int((i + 1) / total_sheets * 100))
                QApplication.processEvents()

            excel.DisplayAlerts = False
            output_workbook.SaveAs(os.path.abspath(save_path))
            excel.DisplayAlerts = True
            output_workbook.Close(SaveChanges=False)

        except Exception as e:
            self.txtLogOutput.append(f"win32 병합 오류: {e}")
        finally:
            if excel:
                excel.DisplayAlerts = False
                excel.Application.Quit()

    def merge_as_sheets(self, sheets_to_merge, save_path):
        output_workbook = openpyxl.Workbook()
        output_workbook.remove(output_workbook.active) # Remove default sheet

        total_sheets = len(sheets_to_merge)
        for i, item in enumerate(sheets_to_merge):
            file_name, sheet_name = item.split('/', 1)
            file_path = self.file_info.get(file_name, {}).get('processed_path')
            
            if not file_path:
                self.txtLogOutput.append(f"파일을 찾을 수 없습니다: {file_name}")
                continue

            try:
                source_workbook = self._open_workbook(file_path, file_name, data_only=self.options['only_value_copy']) # Pass data_only flag
                if not source_workbook:
                    continue

                if file_path.endswith('.xlsx'):
                    source_sheet = source_workbook[sheet_name]
                else: # .xls
                    source_sheet = source_workbook.sheet_by_name(sheet_name)

                # New sheet naming logic
                sheet_name_rule = self.options.get('sheet_name_rule', 'OriginalBoth')
                if sheet_name_rule == 'OriginalBoth':
                    new_sheet_name = f"{os.path.splitext(file_name)[0]}_{sheet_name}"
                else: # OriginalSheet
                    new_sheet_name = sheet_name

                # Sanitize and truncate
                if len(new_sheet_name) > 31:
                    self.txtLogOutput.append(f"시트 이름이 31자를 초과하여 일부 잘립니다: {new_sheet_name}")
                    new_sheet_name = new_sheet_name[:31]
                new_sheet_name = re.sub(r'[\\\[\]\*\:\?/]', '_', new_sheet_name)

                # Handle duplicates
                original_new_sheet_name = new_sheet_name
                counter = 2
                while new_sheet_name in output_workbook.sheetnames:
                    suffix = f" ({counter})"
                    truncated_len = 31 - len(suffix)
                    new_sheet_name = f"{original_new_sheet_name[:truncated_len]}{suffix}"
                    counter += 1
                
                output_sheet = output_workbook.create_sheet(title=new_sheet_name)
                self.copy_sheet_data(source_sheet, output_sheet, file_name=file_name)

            except Exception as e:
                self.txtLogOutput.append(f"시트 복사 오류 {item}: {e}")

            self.progressBar.setValue(int((i + 1) / total_sheets * 100))
            QApplication.processEvents()

        if self.options['only_value_copy']:
            if self.debug_mode:
                self.txtLogOutput.append("DEBUG: 병합된 시트의 수식을 값으로 변환 중 (openpyxl)...")
            for sheet in output_workbook.worksheets:
                for row in sheet.iter_rows():
                    for cell in row:
                        if cell.data_type == 'f':
                            cell.value = cell.value

        output_workbook.save(save_path)

    def merge_horizontally(self, sheets_to_merge, save_path):
        output_workbook = openpyxl.Workbook()
        output_sheet = output_workbook.active
        output_sheet.title = "Merged_Sheet"

        total_sheets = len(sheets_to_merge)
        last_col = 0
        for i, item in enumerate(sheets_to_merge):
            file_name, sheet_name = item.split('/', 1)
            self.lblCurrentFile.setText(f'{item} 병합 중...')

            file_path = self.file_info.get(file_name, {}).get('processed_path')

            if not file_path:
                self.txtLogOutput.append(f"파일을 찾을 수 없습니다: {file_name}")
                continue

            try:
                source_workbook = self._open_workbook(file_path, file_name, data_only=False) # Always load with formulas
                if not source_workbook:
                    continue

                if file_path.endswith('.xlsx'):
                    source_sheet = source_workbook[sheet_name]
                else: # .xls
                    source_sheet = source_workbook.sheet_by_name(sheet_name)
                
                self.copy_sheet_data(source_sheet, output_sheet, start_col=last_col + 1, file_name=file_name)
                last_col = output_sheet.max_column

            except Exception as e:
                self.txtLogOutput.append(f"시트 병합 오류 {item}: {e}")

            self.progressBar.setValue(int((i + 1) / total_sheets * 100))
            QApplication.processEvents()

        if self.options['only_value_copy']:
            self.txtLogOutput.append("수식을 값으로 변환 중...")
            for row in output_sheet.iter_rows():
                for cell in row:
                    if cell.data_type == 'f':
                        cell.value = cell.value

        output_workbook.save(save_path)

    def merge_vertically(self, sheets_to_merge, save_path):
        output_workbook = openpyxl.Workbook()
        output_sheet = output_workbook.active
        output_sheet.title = "Merged_Sheet"

        total_sheets = len(sheets_to_merge)
        last_row = 0
        for i, item in enumerate(sheets_to_merge):
            file_name, sheet_name = item.split('/', 1)
            self.lblCurrentFile.setText(f'{item} 병합 중...')

            file_path = self.file_info.get(file_name, {}).get('processed_path')

            if not file_path:
                self.txtLogOutput.append(f"파일을 찾을 수 없습니다: {file_name}")
                continue

            try:
                source_workbook = self._open_workbook(file_path, file_name, data_only=self.options['only_value_copy']) # Pass data_only flag
                if not source_workbook:
                    continue

                if file_path.endswith('.xlsx'):
                    source_sheet = source_workbook[sheet_name]
                else: # .xls
                    source_sheet = source_workbook.sheet_by_name(sheet_name)

                self.copy_sheet_data(source_sheet, output_sheet, start_row=last_row + 1, file_name=file_name)
                last_row = output_sheet.max_row

            except Exception as e:
                self.txtLogOutput.append(f"시트 병합 오류 {item}: {e}")

            self.progressBar.setValue(int((i + 1) / total_sheets * 100))
            QApplication.processEvents()

        output_workbook.save(save_path)

    def copy_sheet_data(self, source_sheet, output_sheet, start_row=1, start_col=1, file_name=""):
        # SheetTrim logic is simplified and only works for .xlsx (openpyxl)
        if self.options.get('sheet_trim_value', 0) > 0 and isinstance(source_sheet, openpyxl.worksheet.worksheet.Worksheet):
            if self.options.get('sheet_trim_rows', False):
                rows_to_delete = []
                for i, row in enumerate(source_sheet.iter_rows()):
                    if all(cell.value is None or str(cell.value).strip() == "" for cell in row):
                        rows_to_delete.append(i + 1)
                for row_idx in sorted(rows_to_delete, reverse=True):
                    source_sheet.delete_rows(row_idx, 1)

            if self.options.get('sheet_trim_cols', False):
                cols_to_delete = []
                for i, col in enumerate(source_sheet.iter_cols()):
                    if all(cell.value is None or str(cell.value).strip() == "" for cell in col):
                        cols_to_delete.append(i + 1)
                for col_idx in sorted(cols_to_delete, reverse=True):
                    source_sheet.delete_cols(col_idx, 1)

        if isinstance(source_sheet, openpyxl.worksheet.worksheet.Worksheet):
            for row in source_sheet.iter_rows():
                for cell in row:
                    new_cell = output_sheet.cell(row=cell.row + start_row - 1, column=cell.column + start_col - 1)
                    new_cell.value = cell.value
                    if cell.has_style:
                        new_cell.font = cell.font.copy()
                        new_cell.border = cell.border.copy()
                        new_cell.fill = cell.fill.copy()
                        new_cell.number_format = cell.number_format
                        new_cell.protection = cell.protection.copy()
                        new_cell.alignment = cell.alignment.copy()
        elif isinstance(source_sheet, xlrd.sheet.Sheet):
            self.txtLogOutput.append(f".xls 파일({file_name}/{source_sheet.name})의 서식은 일부만 지원됩니다.")
            for row_idx in range(source_sheet.nrows):
                for col_idx in range(source_sheet.ncols):
                    cell_value = source_sheet.cell_value(row_idx, col_idx)
                    output_sheet.cell(row=row_idx + start_row, column=col_idx + start_col).value = cell_value

if __name__ == "__main__":
    app = QApplication(sys.argv)
    window = MainWindow()
    window.show()
    sys.exit(app.exec())
