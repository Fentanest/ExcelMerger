import os
import io
import tempfile
import msoffcrypto
import openpyxl
import xlrd
from dialogs import PasswordDialog
import sys
import csv
from pyxlsb import open_workbook as open_xlsb

class FileHandler:
    def __init__(self, main_window):
        self.main_window = main_window

    def _open_workbook(self, file_path, file_name, data_only=False):
        try:
            lower_path = file_path.lower()
            if lower_path.endswith('.xlsx'):
                return openpyxl.load_workbook(file_path, read_only=False, data_only=data_only)
            elif lower_path.endswith('.xlsm'):
                return openpyxl.load_workbook(file_path, read_only=False, keep_vba=True, data_only=data_only)
            elif lower_path.endswith('.xls'):
                return xlrd.open_workbook(file_path, formatting_info=True)
            elif lower_path.endswith('.xlsb'):
                self.main_window.txtLogOutput.append(f"표준 병합을 위해 .xlsb 파일을 변환 중: {file_name}")
                with open_xlsb(file_path) as wb_xlsb:
                    wb_xlsx = openpyxl.Workbook()
                    wb_xlsx.remove(wb_xlsx.active)
                    for sheet_name in wb_xlsb.sheets:
                        ws_xlsx = wb_xlsx.create_sheet(sheet_name)
                        with wb_xlsb.get_sheet(sheet_name) as sheet_xlsb:
                            for row in sheet_xlsb.rows():
                                ws_xlsx.append([c.v for c in row])
                    return wb_xlsx
            elif lower_path.endswith('.csv'):
                self.main_window.txtLogOutput.append(f"표준 병합을 위해 .csv 파일을 변환 중: {file_name}")
                wb = openpyxl.Workbook()
                ws = wb.active
                ws.title = os.path.splitext(file_name)[0]
                try:
                    with open(file_path, 'r', newline='', encoding='utf-8') as csvfile:
                        reader = csv.reader(csvfile)
                        for row in reader:
                            ws.append(row)
                except UnicodeDecodeError:
                    with open(file_path, 'r', newline='', encoding='cp949') as csvfile:
                        reader = csv.reader(csvfile)
                        for row in reader:
                            ws.append(row)
                return wb
        except Exception as e:
            self.main_window.txtLogOutput.append(f"파일 열기 오류 {file_name}: {e}")
            return None
        return None

    def convert_to_xlsx(self, file_path):
        """Converts various Excel formats to XLSX, returns path to new file."""
        file_name = os.path.basename(file_path)
        lower_file_path = file_path.lower()

        # High-quality conversion for .xls, .xlsb, .xlsm on Windows
        if sys.platform == 'win32' and self.main_window.merger_win32.win32 and lower_file_path.endswith(('.xls', '.xlsb', '.xlsm')):
            return self.main_window.merger_win32.convert_to_xlsx_win32(file_path)

        # CSV to XLSX conversion (all platforms)
        if lower_file_path.endswith('.csv'):
            try:
                self.main_window.txtLogOutput.append(f".csv 파일을 .xlsx로 변환 중: {file_name}")
                wb = openpyxl.Workbook()
                ws = wb.active
                
                try:
                    with open(file_path, 'r', newline='', encoding='utf-8') as csvfile:
                        reader = csv.reader(csvfile)
                        for row in reader:
                            ws.append(row)
                except UnicodeDecodeError:
                    with open(file_path, 'r', newline='', encoding='cp949') as csvfile: # Fallback for Korean
                        reader = csv.reader(csvfile)
                        for row in reader:
                            ws.append(row)

                fd, xlsx_path = tempfile.mkstemp(suffix='.xlsx', prefix='csv_converted_')
                os.close(fd)
                wb.save(xlsx_path)
                self.main_window.temp_files.append(xlsx_path)
                return xlsx_path
            except Exception as e:
                self.main_window.txtLogOutput.append(f".csv to .xlsx 변환 오류: {e}")
                return None

        # Non-Win32 conversions
        try:
            fd, xlsx_path = tempfile.mkstemp(suffix='.xlsx', prefix='converted_')
            os.close(fd)
            
            # .xlsb to .xlsx
            if lower_file_path.endswith('.xlsb'):
                self.main_window.txtLogOutput.append(f".xlsb 파일을 .xlsx로 변환 중: {file_name}")
                with open_xlsb(file_path) as wb_xlsb:
                    wb_xlsx = openpyxl.Workbook()
                    wb_xlsx.remove(wb_xlsx.active)
                    for sheet_name in wb_xlsb.sheets:
                        ws_xlsx = wb_xlsx.create_sheet(sheet_name)
                        with wb_xlsb.get_sheet(sheet_name) as sheet_xlsb:
                            for row in sheet_xlsb.rows():
                                ws_xlsx.append([c.v for c in row])
                    wb_xlsx.save(xlsx_path)
                self.main_window.temp_files.append(xlsx_path)
                return xlsx_path

            # .xlsm to .xlsx
            elif lower_file_path.endswith('.xlsm'):
                self.main_window.txtLogOutput.append(f".xlsm 파일을 .xlsx로 변환 중 (VBA 제외): {file_name}")
                wb = openpyxl.load_workbook(file_path, data_only=True) # data_only to avoid formula issues, removed read_only=True
                wb.save(xlsx_path)
                self.main_window.temp_files.append(xlsx_path)
                return xlsx_path

            # .xls to .xlsx
            elif lower_file_path.endswith('.xls'):
                self.main_window.txtLogOutput.append(f".xls 파일을 .xlsx로 변환 중: {file_name}")
                wb_xls = xlrd.open_workbook(file_path)
                wb_xlsx = openpyxl.Workbook()
                wb_xlsx.remove(wb_xlsx.active)
                for sheet_xls in wb_xls.sheets():
                    ws_xlsx = wb_xlsx.create_sheet(sheet_xls.name)
                    for row in range(sheet_xls.nrows):
                        ws_xlsx.append(sheet_xls.row_values(row))
                wb_xlsx.save(xlsx_path)
                self.main_window.temp_files.append(xlsx_path)
                return xlsx_path

        except Exception as e:
            self.main_window.txtLogOutput.append(f"파일 변환 오류 ({file_name}): {e}")
            return None
        
        return file_path # Should not be reached if conversion was needed

    def get_sheet_names(self, file_path):
        file_name = os.path.basename(file_path)
        processed_file_path = file_path
        file_ext = os.path.splitext(file_name)[1].lower()

        # 1. Handle Encryption
        if file_ext in ['.xlsx', '.xls', '.xlsm', '.xlsb']:
            is_encrypted = False
            try:
                with open(processed_file_path, 'rb') as f:
                    office_file = msoffcrypto.OfficeFile(f)
                    if office_file.is_encrypted():
                        is_encrypted = True
            except Exception:
                pass # Not a valid office file

            if is_encrypted:
                decrypted_temp_path = self.handle_encrypted_file(processed_file_path)
                if decrypted_temp_path:
                    processed_file_path = decrypted_temp_path
                else:
                    self.main_window.txtLogOutput.append(f"파일을 열 수 없습니다 (암호화 문제 또는 사용자 취소): {file_name}")
                    return None, None

        # 2. Get sheet names using Python libraries
        self.main_window.txtLogOutput.append(f"시트 목록 읽기: {file_name}")
        try:
            lower_path = processed_file_path.lower()
            if lower_path.endswith('.xlsm'):
                wb = openpyxl.load_workbook(processed_file_path, read_only=False, keep_vba=True, data_only=True)
                sheet_names = wb.sheetnames
                wb.close()
                return sheet_names, processed_file_path
            elif lower_path.endswith('.xlsx'):
                wb = openpyxl.load_workbook(processed_file_path, read_only=True, data_only=True)
                sheet_names = wb.sheetnames
                wb.close()
                return sheet_names, processed_file_path
            elif lower_path.endswith('.xls'):
                wb = xlrd.open_workbook(processed_file_path, on_demand=True)
                return wb.sheet_names(), processed_file_path
            elif lower_path.endswith('.xlsb'):
                with open_xlsb(processed_file_path) as wb:
                    return wb.sheets, processed_file_path
            elif lower_path.endswith('.csv'):
                return [os.path.splitext(file_name)[0]], processed_file_path
            else:
                self.main_window.txtLogOutput.append(f"지원하지 않는 파일 형식입니다: {file_name}")
                return None, None
                
        except Exception as e:
            self.main_window.txtLogOutput.append(f"시트 이름 가져오기 오류 ({file_name}): {e}")
            return None, None

    def handle_encrypted_file(self, file_path):
        if self.main_window.stop_asking_for_passwords:
            self.main_window.txtLogOutput.append(f"비밀번호 입력을 중단하여 파일을 건너뜁니다: {os.path.basename(file_path)}")
            if self.main_window.debug_mode:
                self.main_window.txtLogOutput.append(f"DEBUG: handle_encrypted_file returning None (stopped asking).")
            return None

        self.main_window.txtLogOutput.append(f"암호화된 파일 감지: {os.path.basename(file_path)}")
        password = None
        decrypted_file_buffer = None
        basename = os.path.basename(file_path)
        temp_decrypted_path = None

        if basename in self.main_window.file_passwords:
            try:
                self.main_window.txtLogOutput.append(f'{basename}에 대해 기억된 비밀번호로 열기 시도...')
                decrypted_file_buffer = io.BytesIO()
                with open(file_path, 'rb') as f:
                    office_file = msoffcrypto.OfficeFile(f)
                    office_file.load_key(password=self.main_window.file_passwords[basename])
                    office_file.decrypt(decrypted_file_buffer)
                password = self.main_window.file_passwords[basename]
                self.main_window.txtLogOutput.append("기억된 비밀번호로 열기 성공.")
            except Exception:
                self.main_window.txtLogOutput.append("기억된 비밀번호 실패.")
                decrypted_file_buffer = None

        if not password and self.main_window.use_global_password and self.main_window.global_password:
            try:
                self.main_window.txtLogOutput.append("전역 비밀번호로 열기 시도...")
                decrypted_file_buffer = io.BytesIO()
                with open(file_path, 'rb') as f:
                    office_file = msoffcrypto.OfficeFile(f)
                    office_file.load_key(password=self.main_window.global_password)
                    office_file.decrypt(decrypted_file_buffer)
                password = self.main_window.global_password
                self.main_window.file_passwords[basename] = self.main_window.global_password
                self.main_window.txtLogOutput.append("전역 비밀번호로 열기 성공.")
            except Exception as e:
                self.main_window.txtLogOutput.append(f"전역 비밀번호 실패: {e}")
                decrypted_file_buffer = None

        if not password:
            if self.main_window.debug_mode:
                self.main_window.txtLogOutput.append(f"DEBUG: Showing PasswordDialog for {basename}.")
            dialog = PasswordDialog(basename, self.main_window)
            dialog.setWindowIcon(self.main_window.windowIcon())
            result = dialog.exec()

            if dialog.stopped:
                self.main_window.stop_asking_for_passwords = True
                self.main_window.txtLogOutput.append("사용자가 중단하여 이후 암호 입력을 건너뜁니다.")
                if self.main_window.debug_mode:
                    self.main_window.txtLogOutput.append(f"DEBUG: handle_encrypted_file returning None (dialog stopped).")
                return None

            if result:
                user_password = dialog.lineEditKeepPassword.text()
                if user_password:
                    try:
                        self.main_window.txtLogOutput.append("사용자 입력 비밀번호로 열기 시도...")
                        decrypted_file_buffer = io.BytesIO()
                        with open(file_path, 'rb') as f:
                            office_file = msoffcrypto.OfficeFile(f)
                            office_file.load_key(password=user_password)
                            office_file.decrypt(decrypted_file_buffer)
                        password = user_password
                        self.main_window.txtLogOutput.append("사용자 입력 비밀번호로 열기 성공.")
                        self.main_window.file_passwords[basename] = user_password
                        if dialog.chkKeepPassword.isChecked():
                            self.main_window.global_password = user_password
                            self.main_window.use_global_password = True
                    except Exception as e:
                        self.main_window.txtLogOutput.append(f"사용자 입력 비밀번호 실패: {e}")
                        if self.main_window.debug_mode:
                            self.main_window.txtLogOutput.append(f"DEBUG: handle_encrypted_file returning None (user password failed).")
                        return None
                else:
                    self.main_window.txtLogOutput.append("비밀번호가 입력되지 않아 파일을 건너뜁니다.")
                    if self.main_window.debug_mode:
                        self.main_window.txtLogOutput.append(f"DEBUG: handle_encrypted_file returning None (empty user password).")
                    return None
            else:
                self.main_window.txtLogOutput.append("사용자가 취소하여 파일을 건너뜁니다.")
                if self.main_window.debug_mode:
                    self.main_window.txtLogOutput.append(f"DEBUG: handle_encrypted_file returning None (dialog cancelled).")
                return None

        if password and decrypted_file_buffer:
            fd, temp_decrypted_path = tempfile.mkstemp(suffix=os.path.splitext(file_path)[1], prefix='decrypted_')
            os.close(fd)
            with open(temp_decrypted_path, 'wb') as tmp_f:
                tmp_f.write(decrypted_file_buffer.getbuffer())
            self.main_window.temp_files.append(temp_decrypted_path)
            if self.main_window.debug_mode:
                self.main_window.txtLogOutput.append(f"DEBUG: handle_encrypted_file returning decrypted path: {temp_decrypted_path}")
            return temp_decrypted_path
        
        if self.main_window.debug_mode:
            self.main_window.txtLogOutput.append(f"DEBUG: handle_encrypted_file returning None (no password worked).")
        return None