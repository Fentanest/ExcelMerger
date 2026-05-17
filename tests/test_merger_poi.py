import base64
import pathlib
import tempfile
import unittest

try:
    from openpyxl import Workbook, load_workbook
except ImportError:  # pragma: no cover - test environment without openpyxl
    Workbook = None
    load_workbook = None

from excelmerger.engines.detector import detect_jpype
from excelmerger.engines.poi import MergerPOI


class _Log:
    def __init__(self):
        self.messages = []

    def append(self, message):
        self.messages.append(message)


class _Bar:
    def setValue(self, value):
        self.value = value


class _Label:
    def setText(self, text):
        self.text = text


class _MainWindowStub:
    def __init__(self, file_info):
        self.file_info = file_info
        self.options = {
            "sheet_name_rule": "OriginalBoth",
            "only_value_copy": False,
            "sheet_trim_value": 0,
            "sheet_trim_rows": False,
            "sheet_trim_cols": False,
        }
        self.txtLogOutput = _Log()
        self.progressBar = _Bar()
        self.lblCurrentFile = _Label()


@unittest.skipUnless(detect_jpype()["available"] and Workbook is not None, "JPype/POI runtime unavailable")
class MergerPOITests(unittest.TestCase):
    def test_merge_as_sheets_preserves_formula(self):
        with tempfile.TemporaryDirectory() as tmpdir:
            source_path = pathlib.Path(tmpdir) / "formula.xlsx"
            output_path = pathlib.Path(tmpdir) / "merged.xlsx"

            workbook = Workbook()
            sheet = workbook.active
            sheet.title = "Sheet1"
            sheet["A1"] = 1
            sheet["A2"] = 2
            sheet["A3"] = "=SUM(A1:A2)"
            workbook.save(source_path)

            file_info = {
                "formula.xlsx": {
                    "processed_path": str(source_path),
                }
            }
            merger = MergerPOI(_MainWindowStub(file_info))
            merger.merge_as_sheets(["formula.xlsx/Sheet1"], str(output_path))

            merged = load_workbook(output_path, data_only=False)
            self.assertEqual("=SUM(A1:A2)", merged["formula_Sheet1"]["A3"].value)
            merged.close()

    def test_merge_as_sheets_preserves_fill_color(self):
        with tempfile.TemporaryDirectory() as tmpdir:
            source_path = pathlib.Path(tmpdir) / "color.xlsx"
            output_path = pathlib.Path(tmpdir) / "merged.xlsx"

            workbook = Workbook()
            sheet = workbook.active
            sheet.title = "Sheet1"
            sheet["A1"] = "color"

            from openpyxl.styles import PatternFill

            sheet["A1"].fill = PatternFill(fill_type="solid", fgColor="DCE6F1")
            workbook.save(source_path)

            file_info = {
                "color.xlsx": {
                    "processed_path": str(source_path),
                }
            }
            merger = MergerPOI(_MainWindowStub(file_info))
            merger.merge_as_sheets(["color.xlsx/Sheet1"], str(output_path))

            merged = load_workbook(output_path, data_only=False)
            color = merged["color_Sheet1"]["A1"].fill.fgColor
            self.assertEqual("rgb", color.type)
            self.assertTrue(color.rgb.endswith("DCE6F1"))
            merged.close()

    def test_merge_as_sheets_preserves_embedded_picture(self):
        png_bytes = base64.b64decode(
            "iVBORw0KGgoAAAANSUhEUgAAAAEAAAABCAQAAAC1HAwCAAAAC0lEQVR42mP8/x8AAwMCAO+X2ioAAAAASUVORK5CYII="
        )

        with tempfile.TemporaryDirectory() as tmpdir:
            source_path = pathlib.Path(tmpdir) / "image.xlsx"
            output_path = pathlib.Path(tmpdir) / "merged.xlsx"

            merger = MergerPOI(_MainWindowStub({}))
            merger._ensure_jvm()

            workbook = merger._XSSFWorkbook()
            try:
                sheet = workbook.createSheet("Sheet1")
                sheet.createRow(0).createCell(0).setCellValue("image")
                drawing = sheet.createDrawingPatriarch()
                anchor = workbook.getCreationHelper().createClientAnchor()
                anchor.setCol1(1)
                anchor.setCol2(2)
                anchor.setRow1(1)
                anchor.setRow2(2)
                picture_index = workbook.addPicture(png_bytes, 6)
                drawing.createPicture(anchor, picture_index)
                merger._save_workbook(workbook, str(source_path))
            finally:
                workbook.close()

            file_info = {
                "image.xlsx": {
                    "processed_path": str(source_path),
                }
            }
            merger = MergerPOI(_MainWindowStub(file_info))
            merger.merge_as_sheets(["image.xlsx/Sheet1"], str(output_path))

            merged = load_workbook(output_path, data_only=False)
            self.assertEqual(1, len(getattr(merged["image_Sheet1"], "_images", [])))
            merged.close()


if __name__ == "__main__":
    unittest.main()
